import sqlite3
import datetime
from flask import Flask, jsonify, request, render_template, redirect, url_for, session
from werkzeug.security import generate_password_hash, check_password_hash
import random
from functools import wraps
import os
import json
import time
from googleapiclient.discovery import build
from google.oauth2.service_account import Credentials as ServiceAccountCredentials
from openpyxl import load_workbook
from google.oauth2.service_account import Credentials

# --- FLASK APP ---

app = Flask(__name__)
app.secret_key = "supersecretkey"  # needed for session

# === DEBUG: Show all attendance status for a given roll number ===
@app.route('/debug/attendance_status/<rollno>', methods=['GET'])
def debug_attendance_status(rollno):
    conn_local = sqlite3.connect('school.db')
    cur = conn_local.cursor()
    cur.execute("SELECT date, status FROM attendance WHERE rollno=? ORDER BY date", (rollno,))
    records = cur.fetchall()
    conn_local.close()
    
    # Calculate attendance stats with the new logic
    def present_status(s):
        if not s or not s.strip():
            return False
    status = s.strip().upper()
    # Normalize some noisy variations
    if status in ('1', 'YES', 'Y', 'PRESENT', 'P'):
        return True
    return False
    
    def absent_status(s):
        if not s or not s.strip():
            return False
    status = s.strip().upper()
    if status in ('0', 'NO', 'N', 'ABSENT', 'A'):
        return True
    return False
    
    def valid_status(s):
        return s and s.strip() != ''
    
    total_days = sum(1 for record in records if valid_status(record[1]))
    present_days = sum(1 for record in records if present_status(record[1]))
    absent_days = sum(1 for record in records if absent_status(record[1]))
    
    return jsonify({
        "rollno": rollno,
        "attendance_records": [
            {"date": r[0], "status": r[1]} for r in records
        ],
        "total_records": len(records),
        "calculated_stats": {
            "total_days": total_days,
            "present_days": present_days,
            "absent_days": absent_days,
            "attendance_percentage": round((present_days / total_days * 100) if total_days > 0 else 0, 2)
        }
    })

# === DEBUG: Show all unique status values in attendance table ===
@app.route('/debug/attendance_statuses', methods=['GET'])
def debug_attendance_statuses():
    conn_local = sqlite3.connect('school.db')
    cur = conn_local.cursor()
    cur.execute("SELECT DISTINCT status FROM attendance WHERE status IS NOT NULL AND status != ''")
    records = cur.fetchall()
    conn_local.close()
    
    unique_statuses = [r[0] for r in records]
    return jsonify({
        "unique_statuses": sorted(unique_statuses),
        "count": len(unique_statuses)
    })

# --- DATABASE SETUP ---
conn = sqlite3.connect('school.db', check_same_thread=False)
c = conn.cursor()
# -------------------------------
# Attendance date resolution
# -------------------------------
def _format_variants(dt: datetime.date) -> list:
    try:
        return [
            dt.strftime('%Y-%m-%d').lower(),
            dt.strftime('%d-%m-%Y').lower(),
            dt.strftime('%d-%b-%Y').lower(),
            dt.strftime('%d-%b-%y').lower(),
        ]
    except Exception:
        # Fallback to today if format fails
        today = datetime.date.today()
        return [
            today.strftime('%Y-%m-%d').lower(),
            today.strftime('%d-%m-%Y').lower(),
            today.strftime('%d-%b-%Y').lower(),
            today.strftime('%d-%b-%y').lower(),
        ]

def _parse_date_maybe(s: str):
    if not s:
        return None
    s = str(s).strip()
    for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d-%b-%Y', '%d-%b-%y'):
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except Exception:
            continue
    return None

def _get_target_date_variants_for_attendance() -> list:
    """Return a list of acceptable string variants for the target date to use when
    checking daily-absent. Preference order:
    1) Today (any accepted format) if the attendance table has any row with today's date
    2) Otherwise, the latest date present in attendance not in the future.
    """
    try:
        # Collect all distinct dates
        c.execute("SELECT DISTINCT LOWER(date) FROM attendance WHERE date IS NOT NULL AND TRIM(date) != ''")
        rows = [r[0] for r in c.fetchall() if r and r[0]]
        if not rows:
            return _format_variants(datetime.date.today())
        today = datetime.date.today()
        today_variants = _format_variants(today)
        # If any row equals today's variant, use today
        if any(r in today_variants for r in rows):
            return today_variants
        # Else choose latest date <= today
        parsed = [(r, _parse_date_maybe(r)) for r in rows]
        parsed_valid = [p for p in parsed if p[1] is not None]
        if not parsed_valid:
            return today_variants
        parsed_valid.sort(key=lambda x: x[1])
        # filter not in future
        not_future = [p for p in parsed_valid if p[1] <= today]
        chosen = (not_future[-1][1] if not_future else parsed_valid[-1][1])
        return _format_variants(chosen)
    except Exception:
        return _format_variants(datetime.date.today())

# Identify if a header looks like a date column (from Google Sheet)
def _is_date_header(label: str) -> bool:
    if not label:
        return False
    s = str(label).strip()
    for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d-%b-%Y', '%d-%b-%y'):
        try:
            datetime.datetime.strptime(s, fmt)
            return True
        except Exception:
            continue
    return False


# Create students table
c.execute('''
CREATE TABLE IF NOT EXISTS students (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    reg_no TEXT,
    rollno TEXT UNIQUE,
    name TEXT,
    dob TEXT,
    gender TEXT,
    aadhar TEXT,
    student_mobile TEXT,
    blood_group TEXT,
    parent_name TEXT,
    parent_mobile TEXT,
    address TEXT,
    nationality TEXT,
    religion TEXT,
    community TEXT,
    caste TEXT,
    day_scholar_or_hosteller TEXT,
    current_semester TEXT,
    seat_type TEXT,
    quota_type TEXT,
    email TEXT,
    pmss TEXT,
    remarks TEXT,
    bus_no TEXT,
    hosteller_room_no TEXT,
    outside_staying_address TEXT,
    owner_ph_no TEXT,
    user_id TEXT UNIQUE,
    password_hash TEXT,
    password_plain TEXT,
    extra_json TEXT
)
''')

# --- DB MIGRATIONS (idempotent) ---
def ensure_teachers_schema():
    required_columns = [
        # 'email', 'phone',  # <-- REMOVE these
        'qualification', 'experience', 'subject', 
        'address', 'date_of_joining', 'salary', 'extra_json', 'role'
    ]
    c.execute("PRAGMA table_info(teachers)")
    existing_columns = [col[1] for col in c.fetchall()]
    
    for col in required_columns:
        if col not in existing_columns:
            c.execute(f"ALTER TABLE teachers ADD COLUMN {col} TEXT")
            print(f"Added column {col} to teachers table")
    
    # Remove email and phone columns if they exist
    # SQLite does not support DROP COLUMN directly, so this is a no-op unless you want to recreate the table.
    # For now, just ignore them in code.

    conn.commit()

def ensure_students_schema():
    required_columns = [
        'reg_no', 'rollno', 'name', 'dob', 'gender', 'aadhar', 'student_mobile', 'blood_group',
        'parent_name', 'parent_mobile', 'address', 'nationality', 'religion', 'community', 'caste',
        'day_scholar_or_hosteller', 'current_semester', 'seat_type', 'quota_type', 'email', 'pmss', 'remarks',
        'bus_no', 'hosteller_room_no', 'outside_staying_address', 'owner_ph_no',
        'user_id', 'password_hash', 'password_plain', 'extra_json'
    ]
    c.execute("PRAGMA table_info(students)")
    existing_columns = [col[1] for col in c.fetchall()]
    for col in required_columns:
        if col not in existing_columns:
            c.execute(f"ALTER TABLE students ADD COLUMN {col} TEXT")
            print(f"Added column {col} to students table")
    conn.commit()

# Create teachers table (must exist before running teacher schema migrations)
c.execute('''
CREATE TABLE IF NOT EXISTS teachers (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    teacher_name TEXT,
    department TEXT,
    user_id TEXT UNIQUE,
    pass_hash TEXT,
    pass_plain TEXT
)
''')

# Create attendance table before any use
c.execute('''
CREATE TABLE IF NOT EXISTS attendance (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    rollno TEXT,
    reg_no TEXT,
    date TEXT,
    status TEXT
)
''')
conn.commit()

# Out passes table + schema ensure
c.execute('''
CREATE TABLE IF NOT EXISTS out_passes (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_role TEXT,
    requester_user_id TEXT,
    requester_name TEXT,
    rollno TEXT,
    department TEXT,
    pass_type TEXT,
    reason TEXT,
    from_datetime TEXT,
    to_datetime TEXT,
    od_duration TEXT,
    od_days INTEGER,
    other_hours TEXT,
    status TEXT DEFAULT 'pending',
    approver_user_id TEXT,
    remarks TEXT,
    -- Two-stage workflow fields
    advisor_status TEXT DEFAULT 'pending',
    hod_status TEXT DEFAULT 'pending',
    advisor_user_id TEXT,
    advisor_remarks TEXT,
    hod_user_id TEXT,
    hod_remarks TEXT,
    created_at INTEGER,
    updated_at INTEGER
)
''')
conn.commit()

def ensure_outpasses_schema():
    required_cols = [
        'advisor_status','hod_status','advisor_user_id','advisor_remarks','hod_user_id','hod_remarks',
        'od_duration','od_days','other_hours'
    ]
    c.execute("PRAGMA table_info(out_passes)")
    existing = {row[1] for row in c.fetchall()}
    for col in required_cols:
        if col not in existing:
            try:
                default_clause = "DEFAULT 'pending'" if col in ('advisor_status','hod_status') else ''
                c.execute(f"ALTER TABLE out_passes ADD COLUMN {col} TEXT {default_clause}")
                conn.commit()
            except Exception as e:
                print(f"ensure_outpasses_schema: failed adding {col}: {e}")

ensure_outpasses_schema()

# Create courses table
c.execute('''
CREATE TABLE IF NOT EXISTS courses (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    course_name TEXT,
    course_code TEXT UNIQUE,
    drive_link TEXT
)
''')
conn.commit()

# Ensure schema up-to-date on startup (after base tables exist)
ensure_teachers_schema()
ensure_students_schema()

# --- Ensure default admin teacher exists ---
def ensure_default_teacher():
    c.execute("SELECT 1 FROM teachers WHERE user_id = ?", ("admin",))
    if not c.fetchone():
        c.execute('''

            INSERT INTO teachers (
                teacher_name, department, user_id, pass_hash, pass_plain, qualification, experience, subject, address, date_of_joining, salary, extra_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            "Admin", "Admin", "admin",
            generate_password_hash("admin123"), "admin123",
            "M.Sc", "10", "All", "Admin Address", "2020-01-01", "0", "{}"
        ))
        conn.commit()

ensure_default_teacher()
# --- HELPER FUNCTIONS ---
def generate_user_id(rollno):
    return f"stu{rollno}"


def generate_password():
    return str(random.randint(100000, 999999))


# --- GOOGLE SHEETS INTEGRATION ---
# Prefer explicit env var, otherwise auto-detect local credentials file names
def _find_credentials_file():
    env_path = os.environ.get("GOOGLE_APPLICATION_CREDENTIALS")
    if env_path and os.path.exists(env_path):
        return env_path
    for candidate in [
        "credentials.json",            # common name you provided
        "service_account.json",        # common alternative
        os.path.join(os.getcwd(), "credentials.json"),
        os.path.join(os.getcwd(), "service_account.json"),
    ]:
        if os.path.exists(candidate):
            return candidate
    return env_path or "credentials.json"

GOOGLE_CREDENTIALS_FILE = _find_credentials_file()
# Defaults wired to provided sheet links; override with env vars if needed
STUDENTS_SHEET_ID = os.environ.get("STUDENTS_SHEET_ID", "11-fZZNhO7MzKaThXLgyqqV_L5gJcjXC9yc7iWlp3fCo")
ATTENDANCE_SHEET_ID = os.environ.get("ATTENDANCE_SHEET_ID", "1OgLsxcweN2xBo1okhKaeCN2D1PGwmCd50kxEFXHohnM")
# Ranges can be either a tab name (entire tab) or A1 range like 'Sheet1!A:F'.
# Default to your actual tab names from the Google Sheets
STUDENTS_RANGE = os.environ.get("STUDENTS_RANGE", "Student_Details!A:AZ")
ATTENDANCE_RANGE = os.environ.get("ATTENDANCE_RANGE", "attendance!A:ZZ")
# Local Excel fallbacks
STUDENTS_XLSX = os.environ.get("STUDENTS_XLSX", os.path.join(os.getcwd(), "students.xlsx"))
ATTENDANCE_XLSX = os.environ.get("ATTENDANCE_XLSX", os.path.join(os.getcwd(), "attendance.xlsx"))
USE_EXCEL_ONLY = os.environ.get("USE_EXCEL_ONLY", "0") in ("1", "true", "True")

# Courses sheet config (from user's link)
COURSES_SHEET_ID = os.environ.get(
    "COURSES_SHEET_ID",
    "1mXifGuP1hGQvgw9SjSTIkyfPAoJRdcrH91C1Wsq9zYo"
)
COURSES_RANGE = os.environ.get("COURSES_RANGE", "Sheet1!A:C")

_sheets_service = None
_last_students_sync_ts = 0
_last_attendance_sync_ts = 0
SYNC_TTL_SECONDS = int(os.environ.get("GSHEETS_SYNC_TTL_SECONDS", "60")) # You can set this environment variable to a lower value if needed

def get_sheets_service():
    global _sheets_service
    if _sheets_service is not None:
        return _sheets_service
    if not os.path.exists(GOOGLE_CREDENTIALS_FILE):
        raise FileNotFoundError(f"Google credentials file not found at {GOOGLE_CREDENTIALS_FILE}")
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets.readonly"
    ]
    # Enforce service-account only to avoid unverified OAuth consent issues
    with open(GOOGLE_CREDENTIALS_FILE, 'r') as f:
        cred_json = json.load(f)
    if not isinstance(cred_json, dict) or cred_json.get('type') != 'service_account':
        raise ValueError(
            "Provided credentials.json is not a service account key. "
            "Create a Service Account JSON in Google Cloud Console and place it as credentials.json, "
            "then share your Google Sheet with the service account's client_email."
        )
    credentials = ServiceAccountCredentials.from_service_account_file(GOOGLE_CREDENTIALS_FILE, scopes=scopes)
    _sheets_service = build('sheets', 'v4', credentials=credentials)
    return _sheets_service

def _split_ids(ids: str):
    # Allow comma-separated multiple spreadsheet IDs
    return [s.strip() for s in str(ids or "").split(',') if s.strip()]

def read_sheet_values(spreadsheet_id, a1_range):
    service = get_sheets_service()
    # If the range does not specify a sheet/tab (no '!'), prefix the first sheet title
    effective_range = a1_range
    try:
        if '!' not in a1_range:
            meta = service.spreadsheets().get(spreadsheetId=spreadsheet_id, fields='sheets(properties(title))').execute()
            sheets = meta.get('sheets', [])
            if not sheets:
                raise ValueError('No sheets found in spreadsheet')
            first_title = sheets[0]['properties']['title']
            effective_range = f"{first_title}!{a1_range}"
        result = service.spreadsheets().values().get(spreadsheetId=spreadsheet_id, range=effective_range).execute()
        return result.get('values', [])
    except Exception as e:
        # Provide clearer hint when the file is not a Google Sheet
        raise RuntimeError(f"Failed to read range '{effective_range}' from spreadsheet '{spreadsheet_id}': {e}")

def read_excel_values(xlsx_path):
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"Excel file not found: {xlsx_path}")
    wb = load_workbook(filename=xlsx_path, data_only=True)
    ws = wb.worksheets[0]
    values = []
    for row in ws.iter_rows(values_only=True):
        values.append(["" if (cell is None) else str(cell) for cell in row])
    # Trim trailing empty rows
    while values and all((v == "" for v in values[-1])):
        values.pop()
    return values


# --- LOADERS (Google Sheets only) ---


def load_students_from_gsheets():
    global _last_students_sync_ts
    if not STUDENTS_SHEET_ID:
        return
    # Merge rows from all provided sheet IDs (first row of the first sheet is treated as headers)
    merged = []
    headers = None
    for sid in _split_ids(STUDENTS_SHEET_ID):
        try:
            vals = read_sheet_values(sid, STUDENTS_RANGE)
        except Exception as e:
            print(f"Error reading student sheet values for sheet ID {sid}: {e}")
            continue
        if not vals:
            print(f"No values returned for student sheet ID {sid}")
            continue
        if headers is None:
            headers = vals[0]
        # append data rows
        merged.extend(vals[1:])
    if not headers:
        print("No student data found in Google Sheet.")
        return
    values = [headers] + merged
    header_map = {str(h).strip().lower(): idx for idx, h in enumerate(headers)}

    def get_by_alias(row, aliases, default=""):
        for alias in aliases:
            idx = header_map.get(alias.lower())
            if idx is not None and idx < len(row) and row[idx] is not None and str(row[idx]).strip() != "":
                return str(row[idx]).strip()
        return default

    # Insert missing students, don't overwrite existing
    for row_idx, row in enumerate(values[1:], start=2):
        try:
            rollno = get_by_alias(row, ['ROLL NO', 'Roll no', 'RollNo', 'rollno'])
            if not rollno:
                continue

            # --- Assign all variables before DB operations ---
            reg_no = get_by_alias(row, ['REG NO', 'Reg no', 'regno'])
            name = get_by_alias(row, ['NAME', 'Name', 'name'])
            dob = get_by_alias(row, ['DOB(DDNOMMNOYYYY)', 'DOB', 'dob', 'Date of Birth'])
            gender = get_by_alias(row, ['GENDER(MALE(or)FEMALE)', 'GENDER', 'Gender', 'gender'])
            aadhar = get_by_alias(row, ['AADHAR(12 DIGITS)', 'AADHAR', 'Aadhar', 'aadhar'])
            student_mobile = get_by_alias(row, [
                'STUDENT MOBILE NUMBER(10 DIGITS)', 'Student Mobile Number', 'student mobile number(10 digits)', 
                'STUDENT MOBILE', 'student_mobile', 'student mobile', 'Phone', 'phone'
            ])
            blood_group = get_by_alias(row, ['BLOOD GROUP', 'Blood Group', 'blood group', 'blood_group'])
            parent_name = get_by_alias(row, ['PARENT/GAURDIAN NAME', 'Parent Name', 'parent name'])
            parent_mobile = get_by_alias(row, [
                'PARENT/GAURDIAN MOBILE NUMBER', 'Parent Mobile', 'parent mobile', 
                'PARENT MOBILE NUMBER', 'parent_mobile', 'parent mobile number'
            ])
            address = get_by_alias(row, ['ADDRESS', 'Address', 'address'])
            nationality = get_by_alias(row, ['NATIONALITY', 'Nationality', 'nationality'])
            religion = get_by_alias(row, ['RELIGION', 'Religion', 'religion'])
            community = get_by_alias(row, ['COMMUNITY', 'Community', 'community'])
            caste = get_by_alias(row, ['CASTE', 'Caste', 'caste'])
            day_scholar_or_hosteller = get_by_alias(row, ['DAYSCHOLAR OR HOSTELLER', 'Day Scholar or Hosteller', 'day scholar or hosteller'])
            current_semester = get_by_alias(row, [
                'DEPARTMENT', 'Department',
                'CURRENT SEMESTER', 'Current Semester', 'current semester'
            ])
            seat_type = get_by_alias(row, ['SEAT TYPE(REGULAR(or)LATERAL)', 'Seat Type', 'seat type'])
            quota_type = get_by_alias(row, ['QUOTA TYPE(GQ(or)MQ)', 'Quota Type', 'quota type'])
            email = get_by_alias(row, ['EMAIL', 'Email', 'email'])
            pmss = get_by_alias(row, ['PMSS (YES/NO)', 'PMSS', 'pmss'])
            remarks = get_by_alias(row, ['REMARKS', 'Remarks', 'remarks'])
            bus_no = get_by_alias(row, [
                'BUS',
                'BUS NO/PRIVATE BUS', 'Bus No', 'bus no', 'BUS NO', 'bus_no', 'Bus Number'
            ])
            hosteller_room_no = get_by_alias(row, ['HOSTELLER ROOM NO.', 'Hosteller Room No', 'hosteller room no'])
            outside_staying_address = get_by_alias(row, [
                'OUTSTAYING  ADDRESS',
                'OUTSTAYING ADDRESS',
                'OUTSIDE STAYING FULL ADDRESS', 'Outside Staying Address', 'outside staying address', 
                'OUTSIDE ADDRESS', 'outside_address', 'Outside Address'
            ])
            owner_ph_no = get_by_alias(row, [
                "OWNER'S PH NO", "Owner's Phone", "owner's phone", "OWNER", "owner_ph_no", "OWNER_PH_NO"
            ])
            user_id = f"stu{rollno}"

            # Capture all remaining fields as extra JSON (header:value mapping)
            extra = {}
            for idx, header in enumerate(headers):
                if idx < len(row):
                    val = row[idx]
                else:
                    val = None
                if header is None or str(header).strip() == "":
                    continue
                extra[str(header)] = None if val is None else str(val)
            try:
                extra_json = json.dumps(extra)
            except Exception:
                extra_json = "{}"
            # --- End assignments ---

            c.execute("SELECT 1 FROM students WHERE rollno=?", (rollno,))
            existing_student = c.fetchone()
            if existing_student:
                # --- UPDATE EXISTING STUDENT (preserve existing password) ---
                c.execute('''
                    UPDATE students SET
                        reg_no=?, name=?, dob=?, gender=?, aadhar=?, student_mobile=?, blood_group=?,
                        parent_name=?, parent_mobile=?, address=?, nationality=?, religion=?, community=?, caste=?,
                        day_scholar_or_hosteller=?, current_semester=?, seat_type=?, quota_type=?, email=?, pmss=?,
                        remarks=?, bus_no=?, hosteller_room_no=?, outside_staying_address=?, owner_ph_no=?,
                        user_id=?, extra_json=?
                    WHERE rollno=?
                ''', (
                    reg_no, name, dob, gender, aadhar, student_mobile, blood_group,
                    parent_name, parent_mobile, address, nationality, religion, community, caste,
                    day_scholar_or_hosteller, current_semester, seat_type, quota_type, email, pmss,
                    remarks, bus_no, hosteller_room_no, outside_staying_address, owner_ph_no,
                    user_id, extra_json, rollno
                ))
                continue
            # --- INSERT NEW STUDENT (generate password only for new students) ---
            password_plain = str(random.randint(100000, 999999))
            password_hash = generate_password_hash(password_plain)
            c.execute('''
                INSERT INTO students (reg_no, rollno, name, dob, gender, aadhar, student_mobile, blood_group, 
                                    parent_name, parent_mobile, address, nationality, religion, community, caste,
                                    day_scholar_or_hosteller, current_semester, seat_type, quota_type, email, 
                                    pmss, remarks, bus_no, hosteller_room_no, outside_staying_address, owner_ph_no,
                                    user_id, password_hash, password_plain, extra_json)
                VALUES (?,?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (reg_no, rollno, name, dob, gender, aadhar, student_mobile, blood_group, 
                  parent_name, parent_mobile, address, nationality, religion, community, caste,
                  day_scholar_or_hosteller, current_semester, seat_type, quota_type, email, 
                  pmss, remarks, bus_no, hosteller_room_no, outside_staying_address, owner_ph_no,
                  user_id, password_hash, password_plain, extra_json))
        except Exception as e:
            print(f"Error processing student row {row_idx}: {e}")
            continue
        quota_type = get_by_alias(row, ['QUOTA TYPE(GQ(or)MQ)', 'Quota Type', 'quota type'])
        email = get_by_alias(row, ['EMAIL', 'Email', 'email'])
        pmss = get_by_alias(row, ['PMSS (YES/NO)', 'PMSS', 'pmss'])
        remarks = get_by_alias(row, ['REMARKS', 'Remarks', 'remarks'])
        bus_no = get_by_alias(row, [
            'BUS',
            'BUS NO/PRIVATE BUS', 'Bus No', 'bus no', 'BUS NO', 'bus_no', 'Bus Number'
        ])
        hosteller_room_no = get_by_alias(row, ['HOSTELLER ROOM NO.', 'Hosteller Room No', 'hosteller room no'])
        outside_staying_address = get_by_alias(row, [
            'OUTSTAYING  ADDRESS',
            'OUTSTAYING ADDRESS',
            'OUTSIDE STAYING FULL ADDRESS', 'Outside Staying Address', 'outside staying address', 
            'OUTSIDE ADDRESS', 'outside_address', 'Outside Address'
        ])
        owner_ph_no = get_by_alias(row, [
            "OWNER'S PH NO", "Owner's Phone", "owner's phone", "OWNER", "owner_ph_no", "OWNER_PH_NO"
        ])
        user_id = f"stu{rollno}"

        # Capture all remaining fields as extra JSON (header:value mapping)
        extra = {}
        for idx, header in enumerate(headers):
            if idx < len(row):
                val = row[idx]
            else:
                val = None
            if header is None or str(header).strip() == "":
                continue
            extra[str(header)] = None if val is None else str(val)
        try:
            extra_json = json.dumps(extra)
        except Exception:
            extra_json = "{}"
        # --- End assignments ---

        c.execute("SELECT 1 FROM students WHERE rollno=?", (rollno,))
        existing_student = c.fetchone()
        if existing_student:
            # --- UPDATE EXISTING STUDENT (preserve existing password) ---
            c.execute('''
                UPDATE students SET
                    reg_no=?, name=?, dob=?, gender=?, aadhar=?, student_mobile=?, blood_group=?,
                    parent_name=?, parent_mobile=?, address=?, nationality=?, religion=?, community=?, caste=?,
                    day_scholar_or_hosteller=?, current_semester=?, seat_type=?, quota_type=?, email=?, pmss=?,
                    remarks=?, bus_no=?, hosteller_room_no=?, outside_staying_address=?, owner_ph_no=?,
                    user_id=?, extra_json=?
                WHERE rollno=?
            ''', (
                reg_no, name, dob, gender, aadhar, student_mobile, blood_group,
                parent_name, parent_mobile, address, nationality, religion, community, caste,
                day_scholar_or_hosteller, current_semester, seat_type, quota_type, email, pmss,
                remarks, bus_no, hosteller_room_no, outside_staying_address, owner_ph_no,
                user_id, extra_json, rollno
            ))
            continue
        # --- INSERT NEW STUDENT (generate password only for new students) ---
        password_plain = str(random.randint(100000, 999999))
        password_hash = generate_password_hash(password_plain)
        
        c.execute('''
            INSERT INTO students (reg_no, rollno, name, dob, gender, aadhar, student_mobile, blood_group, 
                                parent_name, parent_mobile, address, nationality, religion, community, caste,
                                day_scholar_or_hosteller, current_semester, seat_type, quota_type, email, 
                                pmss, remarks, bus_no, hosteller_room_no, outside_staying_address, owner_ph_no,
                                user_id, password_hash, password_plain, extra_json)
            VALUES (?,?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (reg_no, rollno, name, dob, gender, aadhar, student_mobile, blood_group, 
              parent_name, parent_mobile, address, nationality, religion, community, caste,
              day_scholar_or_hosteller, current_semester, seat_type, quota_type, email, 
              pmss, remarks, bus_no, hosteller_room_no, outside_staying_address, owner_ph_no,
              user_id, password_hash, password_plain, extra_json))
    conn.commit()
    _last_students_sync_ts = int(time.time())
    try:
        c.execute("SELECT COUNT(*) FROM students")
        total = c.fetchone()[0]
        print(f"Students import from Google Sheets completed successfully. Total students in DB: {total}")
    except Exception:
        pass

def load_courses_from_gsheets():
    if not COURSES_SHEET_ID:
        return
    merged = []
    headers = None
    for sid in _split_ids(COURSES_SHEET_ID):
        vals = read_sheet_values(sid, COURSES_RANGE)
        if not vals:
            continue
        if headers is None and vals:
            headers = [str(h).strip().lower() for h in vals[0]]
        merged.extend(vals[1:])
    if not headers or not merged:
        print("No course data found in Google Sheet.")
        return
    values = [headers] + merged
    # Determine indices
    def col_idx(*aliases):
        for a in aliases:
            try:
                return headers.index(a.lower())
            except ValueError:
                continue
        return None
    name_idx = col_idx('course name', 'name', 'course')
    code_idx = col_idx('course code', 'code')
    link_idx = col_idx('drive link', 'link', 'url')
    inserted, updated = 0, 0
    for row in values[1:]:
        course_name = (str(row[name_idx]).strip() if name_idx is not None and name_idx < len(row) and row[name_idx] is not None else "")
        course_code = (str(row[code_idx]).strip() if code_idx is not None and code_idx < len(row) and row[code_idx] is not None else "")
        drive_link = (str(row[link_idx]).strip() if link_idx is not None and link_idx < len(row) and row[link_idx] is not None else "")
        if not (course_name or course_code or drive_link):
            continue
        if not course_code and course_name:
            course_code = course_name.replace(" ", "_").upper()
        if not course_code:
            continue
        c.execute("SELECT id FROM courses WHERE course_code=?", (course_code,))
        existing = c.fetchone()
        if existing:
            c.execute(
                "UPDATE courses SET course_name=?, drive_link=? WHERE id=?",
                (course_name, drive_link, existing[0])
            )
            updated += 1
        else:
            c.execute(
                "INSERT INTO courses (course_name, course_code, drive_link) VALUES (?, ?, ?)",
                (course_name, course_code, drive_link)
            )
            inserted += 1
    conn.commit()
    print(f"Courses sync: inserted={inserted}, updated={updated}")

def load_students_from_excel():
    global _last_students_sync_ts
    values = read_excel_values(STUDENTS_XLSX)
    if not values:
        print("No student data found in Excel.")
        return
    headers = [str(h).strip() for h in values[0]]
    header_map = {str(h).strip().lower(): idx for idx, h in enumerate(headers)}

    def get_by_alias(row, aliases, default=""):
        for alias in aliases:
            idx = header_map.get(alias.lower())
            if idx is not None and idx < len(row) and row[idx] is not None and str(row[idx]).strip() != "":
                return str(row[idx]).strip()
        return default

    for row in values[1:]:
        rollno = get_by_alias(row, ['Roll no', 'ROLL NO', 'roll no', 'rollno', 'REG NO', 'reg no', 'regno'])
        if not rollno:
            continue
        c.execute("SELECT 1 FROM students WHERE rollno= ?", (rollno,))
        if c.fetchone():
            continue
        reg_no = get_by_alias(row, ['REG NO', 'reg no', 'regno'])
        name = get_by_alias(row, ['Name', 'NAME', 'name'])
        dob = get_by_alias(row, ['DOB(DDNOMMNOYYYY)', 'DOB', 'dob', 'Date of Birth'])
        gender = get_by_alias(row, ['GENDER(MALE(or)FEMALE)', 'GENDER', 'gender'])
        aadhar = get_by_alias(row, ['AADHAR(12 DIGITS)', 'AADHAR', 'aadhar'])
        student_mobile = get_by_alias(row, [
            'STUDENT MOBILE NUMBER(10 DIGITS)', 'Phone', 'PHONE', 'phone'
        ])
        blood_group = get_by_alias(row, ['BLOOD GROUP', 'Blood Group', 'blood group', 'blood_group'])
        parent_name = get_by_alias(row, ['PARENT/GAURDIAN NAME', 'Parent Name', 'parent name'])
        parent_mobile = get_by_alias(row, [
            'PARENT/GAURDIAN MOBILE NUMBER', 'Parent Mobile', 'parent mobile', 
            'PARENT MOBILE NUMBER', 'parent_mobile', 'parent mobile number'
        ])
        address = get_by_alias(row, ['ADDRESS', 'Address', 'address'])
        nationality = get_by_alias(row, ['NATIONALITY', 'Nationality', 'nationality'])
        religion = get_by_alias(row, ['RELIGION', 'Religion', 'religion'])
        community = get_by_alias(row, ['COMMUNITY', 'Community', 'community'])
        caste = get_by_alias(row, ['CASTE', 'Caste', 'caste'])
        day_scholar_or_hosteller = get_by_alias(row, ['DAYSCHOLAR OR HOSTELLER', 'Day Scholar or Hosteller', 'day scholar or hosteller'])
        current_semester = get_by_alias(row, [
            'DEPARTMENT', 'Department',
            'CURRENT SEMESTER', 'Current Semester', 'current semester'
        ])
        seat_type = get_by_alias(row, ['SEAT TYPE(REGULAR(or)LATERAL)', 'Seat Type', 'seat type'])
        quota_type = get_by_alias(row, ['QUOTA TYPE(GQ(or)MQ)', 'Quota Type', 'quota type'])
        email = get_by_alias(row, ['Email', 'EMAIL', 'email'])
        pmss = get_by_alias(row, ['PMSS (YES/NO)', 'PMSS', 'pmss'])
        remarks = get_by_alias(row, ['REMARKS', 'Remarks', 'remarks'])
        bus_no = get_by_alias(row, [
            'BUS',
            'BUS NO/PRIVATE BUS', 'Bus No', 'bus no', 'BUS NO', 'bus_no', 'Bus Number'
        ])
        hosteller_room_no = get_by_alias(row, ['HOSTELLER ROOM NO.', 'Hosteller Room No', 'hosteller room no'])
        outside_staying_address = get_by_alias(row, [
            'OUTSTAYING  ADDRESS',
            'OUTSTAYING ADDRESS',
            'OUTSIDE STAYING FULL ADDRESS', 'Outside Staying Address', 'outside staying address', 
            'OUTSIDE ADDRESS', 'outside_address', 'Outside Address'
        ])
        owner_ph_no = get_by_alias(row, [
            "OWNER'S PH NO", "Owner's Phone", "owner's phone", "OWNER PH NO", "owner_ph_no", "OWNER_PH_NO", "OWNER"
        ])

        user_id = f"stu{rollno}"

        extra = {}
        for idx, header in enumerate(headers):
            val = row[idx] if idx < len(row) else None
            if header is None or str(header).strip() == "":
                continue
            extra[str(header)] = None if val is None else str(val)
        try:
            extra_json = json.dumps(extra)
        except Exception:
            extra_json = "{}"

        # Check if student already exists
        c.execute("SELECT 1 FROM students WHERE rollno= ?", (rollno,))
        existing_student = c.fetchone()
        if existing_student:
            # --- UPDATE EXISTING STUDENT (preserve existing password) ---
            c.execute('''
                UPDATE students SET
                    reg_no=?, name=?, dob=?, gender=?, aadhar=?, student_mobile=?, blood_group=?,
                    parent_name=?, parent_mobile=?, address=?, nationality=?, religion=?, community=?, caste=?,
                    day_scholar_or_hosteller=?, current_semester=?, seat_type=?, quota_type=?, email=?, pmss=?,
                    remarks=?, bus_no=?, hosteller_room_no=?, outside_staying_address=?, owner_ph_no=?,
                    user_id=?, extra_json=?
                WHERE rollno=?
            ''', (
                reg_no, name, dob, gender, aadhar, student_mobile, blood_group,
                parent_name, parent_mobile, address, nationality, religion, community, caste,
                day_scholar_or_hosteller, current_semester, seat_type, quota_type, email, pmss,
                remarks, bus_no, hosteller_room_no, outside_staying_address, owner_ph_no,
                user_id, extra_json, rollno
            ))
            continue

        # --- INSERT NEW STUDENT (generate password only for new students) ---
        password_plain = str(random.randint(100000, 999999))
        password_hash = generate_password_hash(password_plain)
        
        c.execute('''
            INSERT INTO students (reg_no, rollno, name, dob, gender, aadhar, student_mobile, blood_group, 
                                parent_name, parent_mobile, address, nationality, religion, community, caste,
                                day_scholar_or_hosteller, current_semester, seat_type, quota_type, email, 
                                pmss, remarks, bus_no, hosteller_room_no, outside_staying_address, owner_ph_no,
                                user_id, password_hash, password_plain, extra_json)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ''', (
            reg_no, rollno, name, dob, gender, aadhar, student_mobile, blood_group, 
            parent_name, parent_mobile, address, nationality, religion, community, caste,
            day_scholar_or_hosteller, current_semester, seat_type, quota_type, email, 
            pmss, remarks, bus_no, hosteller_room_no, outside_staying_address, owner_ph_no,
            user_id, password_hash, password_plain, extra_json
        ))
    conn.commit()
    _last_students_sync_ts = int(time.time())
    try:
        c.execute("SELECT COUNT(*) FROM students")
        total = c.fetchone()[0]
        print(f"Students import from Excel completed successfully. Total students in DB: {total}")
    except Exception:
        pass


def load_attendance_from_gsheets():
    global _last_attendance_sync_ts
    print(f"[DEBUG] Starting attendance sync from Google Sheets...")
    print(f"[DEBUG] ATTENDANCE_SHEET_ID: {ATTENDANCE_SHEET_ID}")
    print(f"[DEBUG] ATTENDANCE_RANGE: {ATTENDANCE_RANGE}")
    
    if not ATTENDANCE_SHEET_ID:
        print("[ERROR] No attendance sheet ID configured!")
        return
    
    # NOTE: Do NOT clear the table up-front. Only clear after we have validated
    # that we actually fetched usable data, to avoid wiping existing data on failures.

    merged = []
    headers = None
    for sid in _split_ids(ATTENDANCE_SHEET_ID):
        print(f"[DEBUG] Processing sheet ID: {sid}")
        try:
            vals = read_sheet_values(sid, ATTENDANCE_RANGE)
            print(f"[DEBUG] Retrieved {len(vals) if vals else 0} rows from sheet")
            if not vals:
                print(f"[WARNING] No values returned for sheet ID: {sid}")
                continue
            if headers is None and vals:
                headers = [str(h).strip() for h in vals[0]]
                print(f"[DEBUG] Headers found: {headers[:10]}...")  # Show first 10 headers
            merged.extend(vals[1:])
            print(f"[DEBUG] Added {len(vals[1:])} data rows from sheet {sid}")
        except Exception as e:
            print(f"[ERROR] Failed to read sheet {sid}: {e}")
            continue
    
    if not headers or not merged:
        print("[ERROR] No attendance data found in Google Sheet.")
        print(f"[DEBUG] Headers: {headers}")
        print(f"[DEBUG] Merged data rows: {len(merged)}")
        return
    values = [headers] + merged
    # Expect first columns to include ROLL NO; remaining date columns are those that parse as dates
    rollno_idx = None
    try:
        rollno_idx = next(i for i, h in enumerate(headers) if str(h).strip().lower() in ['roll no', 'rollno', 'roll_no'])
    except StopIteration:
        print("ROLL NO column not found in attendance sheet; trying REG NO fallback")
        try:
            rollno_idx = next(i for i, h in enumerate(headers) if str(h).strip().lower() in ['reg no', 'regno', 'registration no'])
        except StopIteration:
            print("Neither ROLL NO nor REG NO found in attendance sheet")
            return
        return
    date_columns = [i for i, h in enumerate(headers) if i != rollno_idx and _is_date_header(h)]
    print(f"[DEBUG] Attendance import: headers={headers}")
    print(f"[DEBUG] Attendance import: first 3 data rows={merged[:3]}")
    print(f"[DEBUG] Attendance import: rollno_idx={rollno_idx}, date_columns={date_columns}")
    
    # Debug: Show unique status values found in the data
    unique_statuses = set()
    for row in merged[:10]:  # Check first 10 rows for status variety
        for idx in date_columns:
            if idx < len(row) and row[idx] is not None:
                status = str(row[idx]).strip()
                if status:
                    unique_statuses.add(status)
    print(f"[DEBUG] Unique status values found: {sorted(unique_statuses)}")
    # Clear existing attendance data now that we have valid headers and rows
    try:
        c.execute("DELETE FROM attendance")
        conn.commit()
        print("Cleared existing attendance records before inserting new data.")
    except Exception as e:
        print(f"[ERROR] Could not clear attendance table prior to insert: {e}")
        # If we cannot clear safely, abort to avoid mixing old/new rows
        return

    inserted_count = 0
    for row in values[1:]:
        if rollno_idx >= len(row):
            print(f"[DEBUG] Skipping row (rollno_idx out of range): {row}")
            continue
        rollno = str(row[rollno_idx]).strip()
        if not rollno:
            print(f"[DEBUG] Skipping row (no rollno): {row}")
            continue
        for idx in date_columns:
            date_label = str(headers[idx]).strip()
            if not date_label:
                print(f"[DEBUG] Skipping column (no date label): idx={idx}, row={row}")
                continue
            status = str(row[idx]).strip() if idx < len(row) and row[idx] is not None else ''
            # Normalize status to single char P/A when possible
            s_up = status.upper()
            if s_up in ('PRESENT','P','1','YES','Y'):
                status = 'P'
            elif s_up in ('ABSENT','A','0','NO','N'):
                status = 'A'
            c.execute("INSERT INTO attendance (rollno, date, status) VALUES (?, ?, ?)", (rollno, date_label, status))
            inserted_count += 1
    print(f"[DEBUG] Attendance import: total rows inserted={inserted_count}")
    conn.commit()
    _last_attendance_sync_ts = int(time.time())
    
    # Verify the data was actually inserted
    try:
        c.execute("SELECT COUNT(*) FROM attendance")
        total = c.fetchone()[0]
        print(f"Attendance import from Google Sheets completed successfully. Total attendance rows in DB: {total}")
        
        # Show sample of inserted data
        c.execute("SELECT rollno, date, status FROM attendance LIMIT 5")
        sample_records = c.fetchall()
        print(f"[DEBUG] Sample attendance records: {sample_records}")
        
        # Check for any students with attendance data
        c.execute("SELECT DISTINCT rollno FROM attendance LIMIT 10")
        students_with_attendance = c.fetchall()
        print(f"[DEBUG] Students with attendance data: {[s[0] for s in students_with_attendance]}")
        
    except Exception as e:
        print(f"[ERROR] Failed to verify attendance data: {e}")
        print("Attendance import from Google Sheets completed with errors.")

def load_attendance_from_excel():
    global _last_attendance_sync_ts
    
    # Clear existing attendance data to prevent duplicates
    try:
        c.execute("DELETE FROM attendance")
        conn.commit()
        print("Cleared existing attendance records before syncing from Excel.")
    except Exception as e:
        print(f"Error clearing attendance table: {e}")
        return

    values = read_excel_values(ATTENDANCE_XLSX)
    if not values or len(values) < 2:
        print("No attendance data found in Excel.")
        return
    headers = [str(h).strip() for h in values[0]]
    rollno_idx = None
    try:
        rollno_idx = next(i for i, h in enumerate(headers) if str(h).strip().lower() in ['roll no', 'rollno', 'roll_no'])
    except StopIteration:
        try:
            rollno_idx = next(i for i, h in enumerate(headers) if str(h).strip().lower() in ['reg no', 'regno', 'registration no'])
        except StopIteration:
            print("Neither ROLL NO nor REG NO found in attendance Excel")
            return
    date_columns = [i for i, h in enumerate(headers) if i != rollno_idx]
    for row in values[1:]:
        if rollno_idx >= len(row):
            continue
        rollno = str(row[rollno_idx]).strip()
        if not rollno:
            continue
        for idx in date_columns:
            date_label = str(headers[idx]).strip()
            if not date_label:
                continue
            # Always insert, even if blank or other value
            status = str(row[idx]).strip() if idx < len(row) and row[idx] is not None else ''
            c.execute("INSERT INTO attendance (rollno, date, status) VALUES (?, ?, ?)", (rollno, date_label, status))
    conn.commit()
    _last_attendance_sync_ts = int(time.time())
    try:
        c.execute("SELECT COUNT(*) FROM attendance")
        total = c.fetchone()[0]
        print(f"Attendance import from Excel completed successfully. Total attendance rows in DB: {total}")
    except Exception:
        print("Attendance import from Excel completed successfully.")

# Initial data load: Google Sheets only
# Note: Passwords are only generated for NEW students, existing students keep their current passwords
try:
    if not USE_EXCEL_ONLY and STUDENTS_SHEET_ID:
        load_students_from_gsheets()
except Exception as e:
    print("Error loading students:", e)
    try:
        if os.path.exists(STUDENTS_XLSX):
            load_students_from_excel()
    except Exception as e2:
        print("Excel load students failed:", e2)

try:
    if not USE_EXCEL_ONLY and ATTENDANCE_SHEET_ID:
        load_attendance_from_gsheets()
except Exception as e:
    print("Error loading attendance:", e)
    try:
        if os.path.exists(ATTENDANCE_XLSX):
            load_attendance_from_excel()
    except Exception as e2:
        print("Excel load attendance failed:", e2)

# Try to load courses at startup (no Excel fallback defined)
try:
    if not USE_EXCEL_ONLY and COURSES_SHEET_ID:
        load_courses_from_gsheets()
except Exception as e:
    print("Error loading courses:", e)

def login_required(role):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user' not in session or session.get('role') != role:
                # If the request prefers JSON (fetch/XHR), return 401 JSON instead of HTML redirect
                wants_json = 'application/json' in request.headers.get('Accept', '') or request.headers.get('X-Requested-With') == 'XMLHttpRequest'
                if wants_json:
                    return jsonify({"success": False, "message": "Unauthorized"}), 401
                return redirect(url_for('home'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def login_required_any(roles):
    roles = tuple((r or '').lower() for r in (roles or ()))
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            current_role = (session.get('role') or '').lower()
            if 'user' not in session or current_role not in roles:
                wants_json = 'application/json' in request.headers.get('Accept', '') or request.headers.get('X-Requested-With') == 'XMLHttpRequest'
                if wants_json:
                    return jsonify({"success": False, "message": "Unauthorized"}), 401
                return redirect(url_for('home'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# --- DEFAULT ADMIN CREDENTIALS ---
admin_credentials = {
    "username": "admin",
    "password": generate_password_hash("admin123")
}

# === MANUAL ATTENDANCE SYNC ENDPOINT ===
@app.route('/sync_attendance', methods=['POST'])
@login_required('admin')
def manual_sync_attendance():
    """Manually trigger attendance sync from Google Sheets"""
    try:
        print("[MANUAL SYNC] Starting manual attendance sync...")
        load_attendance_from_gsheets()
        
        # Check how many records were inserted
        conn_local = sqlite3.connect('school.db')
        cur = conn_local.cursor()
        cur.execute("SELECT COUNT(*) FROM attendance")
        total_records = cur.fetchone()[0]
        conn_local.close()
        
        return jsonify({
            "success": True,
            "message": f"Attendance sync completed. Total records in database: {total_records}",
            "total_records": total_records
        })
    except Exception as e:
        print(f"[MANUAL SYNC ERROR] {e}")
        return jsonify({
            "success": False,
            "message": f"Attendance sync failed: {str(e)}"
        }), 500

# === COMPREHENSIVE ATTENDANCE TEST ENDPOINT ===
@app.route('/test_attendance_connection', methods=['GET'])
@login_required('admin')
def test_attendance_connection():
    """Test Google Sheets connection and show detailed information"""
    results = {
        "success": False,
        "steps": [],
        "errors": [],
        "data_preview": None
    }
    
    # Step 1: Check credentials
    try:
        if not os.path.exists(GOOGLE_CREDENTIALS_FILE):
            results["errors"].append(f"Credentials file not found: {GOOGLE_CREDENTIALS_FILE}")
            return jsonify(results)
        
        with open(GOOGLE_CREDENTIALS_FILE, 'r') as f:
            cred_data = json.load(f)
        
        if cred_data.get('type') != 'service_account':
            results["errors"].append("Credentials file is not a service account")
            return jsonify(results)
        
        results["steps"].append("✓ Credentials file found and valid")
        results["service_account_email"] = cred_data.get('client_email')
        
    except Exception as e:
        results["errors"].append(f"Error reading credentials: {e}")
        return jsonify(results)
    
    # Step 2: Test Google Sheets service
    try:
        service = get_sheets_service()
        results["steps"].append("✓ Google Sheets service initialized")
    except Exception as e:
        results["errors"].append(f"Error initializing Google Sheets service: {e}")
        return jsonify(results)
    
    # Step 3: Test reading the attendance sheet
    try:
        print(f"[TEST] Attempting to read sheet {ATTENDANCE_SHEET_ID} with range {ATTENDANCE_RANGE}")
        values = read_sheet_values(ATTENDANCE_SHEET_ID, ATTENDANCE_RANGE)
        
        if not values:
            results["errors"].append("No data returned from Google Sheet")
            return jsonify(results)
        
        results["steps"].append(f"✓ Successfully read {len(values)} rows from Google Sheet")
        results["data_preview"] = {
            "total_rows": len(values),
            "headers": values[0] if values else [],
            "sample_data": values[1:6] if len(values) > 1 else []  # First 5 data rows
        }
        
        # Step 4: Analyze the data structure
        headers = values[0] if values else []
        rollno_idx = None
        try:
            rollno_idx = next(i for i, h in enumerate(headers) if str(h).strip().lower() in ['roll no', 'rollno', 'roll_no'])
            results["steps"].append(f"✓ Found ROLL NO column at index {rollno_idx}")
        except StopIteration:
            results["errors"].append("ROLL NO column not found in headers")
            return jsonify(results)
        
        # Step 5: Check for attendance data
        date_columns = [i for i, h in enumerate(headers) if i != rollno_idx]
        results["steps"].append(f"✓ Found {len(date_columns)} date columns")
        
        # Count unique status values
        unique_statuses = set()
        for row in values[1:11]:  # Check first 10 data rows
            for idx in date_columns:
                if idx < len(row) and row[idx] is not None:
                    status = str(row[idx]).strip()
                    if status:
                        unique_statuses.add(status)
        
        results["unique_statuses"] = sorted(list(unique_statuses))
        results["steps"].append(f"✓ Found unique status values: {results['unique_statuses']}")
        
        results["success"] = True
        
    except Exception as e:
        results["errors"].append(f"Error reading attendance sheet: {e}")
        return jsonify(results)
    
    return jsonify(results)

# === ADMIN DEBUG DASHBOARD ===
@app.route('/admin_debug')
@login_required('admin')
def admin_debug():
    """Admin debug dashboard for testing attendance sync"""
    return '''
    <!DOCTYPE html>
    <html>
    <head>
        <title>Admin Debug Dashboard</title>
        <style>
            body { font-family: Arial, sans-serif; margin: 20px; }
            .section { margin: 20px 0; padding: 15px; border: 1px solid #ccc; border-radius: 5px; }
            button { padding: 10px 15px; margin: 5px; background: #007bff; color: white; border: none; border-radius: 3px; cursor: pointer; }
            button:hover { background: #0056b3; }
            .result { margin: 10px 0; padding: 10px; background: #f8f9fa; border-radius: 3px; }
            .error { background: #f8d7da; color: #721c24; }
            .success { background: #d4edda; color: #155724; }
            pre { background: #f8f9fa; padding: 10px; border-radius: 3px; overflow-x: auto; }
        </style>
    </head>
    <body>
        <h1>Admin Debug Dashboard</h1>
        
        <div class="section">
            <h3>System Health Check</h3>
            <button onclick="checkHealth()">Check System Health</button>
            <div id="health-result" class="result"></div>
        </div>
        
        <div class="section">
            <h3>Google Sheets Connection Test</h3>
            <button onclick="testConnection()">Test Google Sheets Connection</button>
            <div id="connection-result" class="result"></div>
        </div>
        
        <div class="section">
            <h3>Attendance Sync</h3>
            <button onclick="syncAttendance()">Sync Attendance from Google Sheets</button>
            <div id="sync-result" class="result"></div>
        </div>
        
        <div class="section">
            <h3>Database Status</h3>
            <button onclick="checkDatabase()">Check Database Status</button>
            <div id="database-result" class="result"></div>
        </div>
        
        <script>
            async function checkHealth() {
                try {
                    const response = await fetch('/health');
                    const data = await response.json();
                    document.getElementById('health-result').innerHTML = 
                        '<pre>' + JSON.stringify(data, null, 2) + '</pre>';
                } catch (error) {
                    document.getElementById('health-result').innerHTML = 
                        '<div class="error">Error: ' + error.message + '</div>';
                }
            }
            
            async function testConnection() {
                try {
                    const response = await fetch('/test_attendance_connection');
                    const data = await response.json();
                    const resultDiv = document.getElementById('connection-result');
                    if (data.success) {
                        resultDiv.innerHTML = '<div class="success">Connection successful!</div><pre>' + 
                            JSON.stringify(data, null, 2) + '</pre>';
                    } else {
                        resultDiv.innerHTML = '<div class="error">Connection failed!</div><pre>' + 
                            JSON.stringify(data, null, 2) + '</pre>';
                    }
                } catch (error) {
                    document.getElementById('connection-result').innerHTML = 
                        '<div class="error">Error: ' + error.message + '</div>';
                }
            }
            
            async function syncAttendance() {
                try {
                    const response = await fetch('/sync_attendance', { method: 'POST' });
                    const data = await response.json();
                    const resultDiv = document.getElementById('sync-result');
                    if (data.success) {
                        resultDiv.innerHTML = '<div class="success">Sync successful!</div><pre>' + 
                            JSON.stringify(data, null, 2) + '</pre>';
                    } else {
                        resultDiv.innerHTML = '<div class="error">Sync failed!</div><pre>' + 
                            JSON.stringify(data, null, 2) + '</pre>';
                    }
                } catch (error) {
                    document.getElementById('sync-result').innerHTML = 
                        '<div class="error">Error: ' + error.message + '</div>';
                }
            }
            
            async function checkDatabase() {
                try {
                    const response = await fetch('/debug/attendance_statuses');
                    const data = await response.json();
                    document.getElementById('database-result').innerHTML = 
                        '<pre>' + JSON.stringify(data, null, 2) + '</pre>';
                } catch (error) {
                    document.getElementById('database-result').innerHTML = 
                        '<div class="error">Error: ' + error.message + '</div>';
                }
            }
        </script>
    </body>
    </html>
    '''


# ====================================================
# ROUTES
# ====================================================
@app.route('/')
def home():
    return render_template('index.html')



@app.route('/current_student_info')
@login_required('student')
def current_student_info():
    user_id = session.get('user')
    if not user_id:
        return jsonify({"success": False, "message": "Not logged in"}), 401
    
    conn = sqlite3.connect("school.db")
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT name, rollno, current_semester FROM students WHERE user_id=?", (user_id,))
    student = cur.fetchone()
    conn.close()
    
    if student:
        return jsonify({"success": True, "student": dict(student)})
    return jsonify({"success": False, "message": "Student not found"}), 404

@app.route('/student_details')
@login_required('student')
def student_details():
    user_id = session.get('user')
    if not user_id:
        return jsonify({"error": "Not logged in"}), 401

    conn = sqlite3.connect("school.db")
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT * FROM students WHERE user_id=?", (user_id,))
    student_row = cur.fetchone()
    conn.close()

    if not student_row:
        return jsonify({"error": "Student not found"}), 404

    student_dict = dict(student_row)
    # Remove sensitive data before sending
    student_dict.pop('password_hash', None)
    student_dict.pop('password_plain', None)
    return jsonify(student_dict)


@app.route('/student_login', methods=['POST'])
def student_login():
    username = request.form.get('username')
    password = request.form.get('password')
    c.execute("SELECT user_id, password_hash FROM students WHERE user_id=?", (username,))
    user = c.fetchone()
    if user and check_password_hash(user[1], password):
        session['user'] = username
        session['role'] = 'student'
        return redirect(url_for('student_dashboard'))
    return render_template('index.html', error="Invalid student credentials")


@app.route('/staff_login', methods=['POST'])
def staff_login():
    username = request.form.get('username')
    password = request.form.get('password')

    if username == "admin" and check_password_hash(admin_credentials["password"], password):
        session['user'] = username
        session['role'] = 'admin'
        return redirect(url_for('admin_dashboard'))

    c.execute("SELECT user_id, pass_hash, COALESCE(role,'') as role FROM teachers WHERE user_id=?", (username,))
    teacher = c.fetchone()
    if teacher and check_password_hash(teacher[1], password):
        role = (teacher[2] or '').strip().lower() or 'teacher'
        session['user'] = username
        session['role'] = role
        if role == 'hod':
            return redirect(url_for('hod_dashboard'))
        if role == 'principal':
            return redirect(url_for('principal_dashboard'))
        return redirect(url_for('teacher_dashboard'))

    return render_template('index.html', error="Invalid staff credentials")


@app.route('/admin_dashboard')
@login_required('admin')
def admin_dashboard():
    return render_template('admin_dashboard.html')


@app.route('/teacher_dashboard')
@login_required('teacher')
def teacher_dashboard():
    return render_template('teacher_dashboard.html')

@app.route('/hod_dashboard')
@login_required('hod')
def hod_dashboard():
    # Fetch HOD's department
    try:
        c.execute("SELECT department FROM teachers WHERE user_id=?", (session.get('user'),))
        row = c.fetchone()
        department = (row[0] or '').strip() if row and row[0] else None
    except Exception:
        department = None

    # Fetch department students
    if department:
        c.execute("SELECT * FROM students WHERE current_semester=?", (department,))
    else:
        c.execute("SELECT * FROM students")
    students = c.fetchall()
    columns = [desc[0] for desc in c.description]

    # Process student data
    department_students = []
    hostellers = []
    day_scholars = []
    outstaying_students = []

    for student in students:
        student_dict = dict(zip(columns, student))
        # Parse extra_json if present
        extra = {}
        if student_dict.get("extra_json"):
            try:
                extra = json.loads(student_dict["extra_json"])
            except Exception:
                extra = {}
        # Merge extra fields, but don't overwrite main columns
        for k, v in extra.items():
            if k not in student_dict or not student_dict[k]:
                student_dict[k] = v
        # Remove sensitive fields
        student_dict.pop("password_hash", None)
        student_dict.pop("password_plain", None)
        student_dict.pop("extra_json", None)

        # Categorize students
        department_students.append(student_dict)
        if student_dict.get("day_scholar_or_hosteller", "").lower() == "hosteller":
            hostellers.append(student_dict)
        elif student_dict.get("day_scholar_or_hosteller", "").lower() == "day scholar":
            day_scholars.append(student_dict)
        if student_dict.get("outside_staying_address"):
            outstaying_students.append(student_dict)

    # Fetch courses
    c.execute("SELECT * FROM courses")
    courses = c.fetchall()

    # Fetch attendance data
    c.execute("SELECT * FROM attendance")
    attendance = c.fetchall()

    return render_template(
        'hod_dashboard.html',
        department_students=department_students,
        hostellers=hostellers,
        day_scholars=day_scholars,
        outstaying_students=outstaying_students,
        courses=courses,
        attendance=attendance
    )

@app.route('/principal_dashboard')
@login_required('principal')
def principal_dashboard():
    return render_template('principal_dashboard.html')

@app.route('/student_dashboard')
@login_required('student')
def student_dashboard():
    return render_template('student_dashboard.html')



@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('home'))


# ====================================================
# API ENDPOINTS
# ====================================================
@app.route('/students', methods=['GET'])
def get_students():
    conn = sqlite3.connect("school.db")
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    # Role-based filtering
    role = session.get('role')
    dept = None
    if role in ('teacher', 'hod'):
        try:
            cur.execute("SELECT department FROM teachers WHERE user_id=?", (session.get('user'),))
            row = cur.fetchone()
            dept = (row[0] or '').strip() if row and row[0] else None
        except Exception:
            dept = None

    # Optional department filter via query param
    q_dept = (request.args.get('dept') or '').strip()

    if role in ('admin', 'principal'):
        # Admin/Principal can view any department
        if q_dept:
            cur.execute("SELECT * FROM students WHERE current_semester = ?", (q_dept,))
        else:
            cur.execute("SELECT * FROM students")
    elif role == 'hod':
        # HOD can also query any department via ?dept=, default to own department
        if q_dept:
            cur.execute("SELECT * FROM students WHERE current_semester = ?", (q_dept,))
        elif dept:
            cur.execute("SELECT * FROM students WHERE current_semester = ?", (dept,))
        else:
            cur.execute("SELECT * FROM students")
    else:
        # Teacher remains restricted to their department
        if dept:
            cur.execute("SELECT * FROM students WHERE current_semester = ?", (dept,))
        else:
            cur.execute("SELECT * FROM students")
    rows = cur.fetchall()
    conn.close()

    student_list = []
    for s in rows:
        # Parse extra_json if present
        extra = {}
        if "extra_json" in s.keys() and s["extra_json"]:
            try:
                extra = json.loads(s["extra_json"])
            except Exception:
                extra = {}

        # Helper → pick from DB, else from extra_json
        def safe_get(col, fallback=""):
            val = s[col] if col in s.keys() else None
            if val is None or str(val).strip() == "":
                return extra.get(col, fallback)
            return val

        # Alias-aware getter for fields that may come under different header names in extra_json
        def get_with_alias(primary_key, aliases, fallback=""):
            # 1) Prefer DB column value if present
            if primary_key in s.keys():
                db_val = s[primary_key]
                if db_val is not None and str(db_val).strip() != "":
                    return db_val
            # 2) Try exact primary key inside extra_json
            if primary_key in extra and str(extra.get(primary_key, "")).strip() != "":
                return extra.get(primary_key)
            # 3) Try aliases inside extra_json
            for k in aliases:
                v = extra.get(k)
                if v is not None and str(v).strip() != "":
                    return v
            return fallback

        # Normalize common fields coming from Google Sheet headers
        owner_value = get_with_alias(
            "owner_ph_no",
            [
                "OWNER'S PH NO",
                "OWNER PH NO",
                "Owner's Phone",
                "owner's phone",
                "OWNER_PH_NO",
                "OWNER"
            ]
        )
        reg_no_value = get_with_alias("reg_no", ["REG NO", "Reg No", "RegNo"]) or safe_get("reg_no")
        rollno_value = get_with_alias("rollno", ["ROLL NO", "Roll No", "RollNo"]) or safe_get("rollno")
        name_value = get_with_alias("name", ["NAME"]) or safe_get("name")
        dob_value = get_with_alias("dob", ["DOB"]) or safe_get("dob")
        gender_value = get_with_alias("gender", ["GENDER"]) or safe_get("gender")
        aadhar_value = get_with_alias("aadhar", ["AADHAR", "AADHAAR"]) or safe_get("aadhar")
        student_mobile_value = get_with_alias("student_mobile", ["STUDENT MOBILE", "STUDENT MOBILE NUMBER", "STUDENT PHONE"]) or safe_get("student_mobile")
        blood_group_value = get_with_alias("blood_group", ["BLOOD GROUP", "BLOODGROUP"]) or safe_get("blood_group")
        parent_name_value = get_with_alias("parent_name", ["PARENT NAME", "FATHER NAME", "GUARDIAN NAME"]) or safe_get("parent_name")
        parent_mobile_value = get_with_alias("parent_mobile", ["PARENT MOBILE NUMBER", "PARENT MOBILE", "PARENT PHONE"]) or safe_get("parent_mobile")
        address_value = get_with_alias("address", ["ADDRESS"]) or safe_get("address")
        nationality_value = get_with_alias("nationality", ["NATIONALITY"]) or safe_get("nationality")
        religion_value = get_with_alias("religion", ["RELIGION"]) or safe_get("religion")
        community_value = get_with_alias("community", ["COMMUNITY", "Community"]) or safe_get("community")
        caste_value = get_with_alias("caste", ["CASTE"]) or safe_get("caste")
        dsh_value = get_with_alias("day_scholar_or_hosteller", ["DAYSCHOLAR OR HOSTELLER", "DAY SCHOLAR OR HOSTELLER"]) or safe_get("day_scholar_or_hosteller")
        department_value = get_with_alias("department", ["DEPARTMENT"])  # separate from current_semester if present
        current_semester_value = get_with_alias("current_semester", ["CURRENT SEMESTER", "CLASS", "SECTION", "SEMESTER"]) or safe_get("current_semester")
        seat_type_value = get_with_alias("seat_type", ["SEAT TYPE"]) or safe_get("seat_type")
        quota_type_value = get_with_alias("quota_type", ["QUOTA TYPE"]) or safe_get("quota_type")
        email_value = get_with_alias("email", ["EMAIL", "Email"]) or safe_get("email")
        pmss_value = get_with_alias("pmss", ["PMSS"]) or safe_get("pmss")
        scholarship_value = get_with_alias("scholarship", ["SCHOLARSHIP", "Scholarship"]) or safe_get("scholarship")
        bus_no_value = get_with_alias("bus_no", ["BUS", "BUS NO", "BUS NUMBER"]) or safe_get("bus_no")
        hosteller_room_no_value = get_with_alias("hosteller_room_no", ["HOSTELLER ROOM NO", "HOSTEL ROOM NO", "ROOM NO"]) or safe_get("hosteller_room_no")
        outside_addr_value = get_with_alias(
            "outside_staying_address",
            [
                "OUTSTAYING  ADDRESS",
                "OUTSTAYING ADDRESS",
                "OUTSIDE STAYING ADDRESS",
                "OUT-STAYING ADDRESS"
            ]
        ) or safe_get("outside_staying_address")

        student_data = {
            "id": s["id"],
            "reg_no": reg_no_value,
            "rollno": rollno_value,
            "name": name_value,
            "dob": dob_value,
            "gender": gender_value,
            "aadhar": aadhar_value,
            "student_mobile": student_mobile_value,
            "blood_group": blood_group_value,
            "parent_name": parent_name_value,
            "parent_mobile": parent_mobile_value,
            "address": address_value,
            "nationality": nationality_value,
            "religion": religion_value,
            "community": community_value,
            "caste": caste_value,
            "day_scholar_or_hosteller": dsh_value,
            "department": department_value,
            "current_semester": current_semester_value,
            "seat_type": seat_type_value,
            "quota_type": quota_type_value,
            "email": email_value,
            "pmss": pmss_value,
            "scholarship": scholarship_value,
            "remarks": safe_get("remarks"),
            "bus_no": bus_no_value,
            "hosteller_room_no": hosteller_room_no_value,
            "outside_staying_address": outside_addr_value,
            "owner_ph_no": owner_value,
            # Convenience duplicates for frontend variations
            "owner": owner_value,
            "owner_phone": owner_value,
            "user_id": safe_get("user_id"),
            "password": s["password_plain"] or "",
            "extra": extra,
            # Convenience aliases for frontend
            "class": current_semester_value,
            "phone": student_mobile_value
        }
        student_list.append(student_data)

    return jsonify(student_list)

@app.route('/departments', methods=['GET'])
def list_departments():
    """Return distinct department values (from students.current_semester)."""
    conn = sqlite3.connect("school.db")
    cur = conn.cursor()
    cur.execute("SELECT DISTINCT current_semester FROM students WHERE current_semester IS NOT NULL AND TRIM(current_semester) != '' ORDER BY current_semester")
    rows = cur.fetchall()
    conn.close()
    return jsonify([r[0] for r in rows])

@app.route('/courses', methods=['GET'])
def get_courses():
    conn_local = sqlite3.connect('school.db')
    conn_local.row_factory = sqlite3.Row
    cur = conn_local.cursor()
    cur.execute("SELECT id, course_name, course_code, drive_link FROM courses ORDER BY course_name")
    rows = cur.fetchall()
    # Auto-sync from Google Sheets if table is empty
    if not rows:
        try:
            if not USE_EXCEL_ONLY and COURSES_SHEET_ID:
                load_courses_from_gsheets()
                cur.execute("SELECT id, course_name, course_code, drive_link FROM courses ORDER BY course_name")
                rows = cur.fetchall()
        except Exception as e:
            print("Auto-sync courses failed:", e)
    conn_local.close()
    return jsonify([
        {
            'id': r['id'],
            'course_name': r['course_name'] or '',
            'course_code': r['course_code'] or '',
            'drive_link': r['drive_link'] or ''
        } for r in rows
    ])

# =========================
# Out Pass APIs
# =========================

PASS_TYPES = (
    'out_pass',           # going out of campus (single day)
    'od_pass',            # on duty (half/full/multi-days)
    'emergency',          # emergency (single day)
    'other'               # hours based / specific time window
)

def _now_epoch() -> int:
    try:
        return int(time.time())
    except Exception:
        return 0

@app.route('/out_pass', methods=['POST'])
@login_required('student')
def create_out_pass():
    data = request.get_json() or {}
    pass_type = (data.get('pass_type') or '').strip().lower()
    reason = (data.get('reason') or '').strip()
    from_datetime = (data.get('from_datetime') or '').strip()
    to_datetime = (data.get('to_datetime') or '').strip()
    # Optional fields for new flows
    od_duration = (data.get('od_duration') or '').strip()  # 'half_day' | 'full_day' | 'n_days'
    od_days = int(data.get('od_days') or 0)
    other_hours = (data.get('other_hours') or '').strip()

    if pass_type not in PASS_TYPES:
        return jsonify({'success': False, 'message': 'Invalid pass type'}), 400
    # Validate based on type
    if pass_type == 'out_pass':
        # Out pass: reason only; from/to optional
        pass
    elif pass_type == 'emergency':
        # Emergency: reason only; from/to optional
        pass
    elif pass_type == 'od_pass':
        # OD: half_day/full_day/n_days; still requires a from/to window
        if od_duration not in ('half_day', 'full_day', 'n_days'):
            return jsonify({'success': False, 'message': 'Invalid OD duration'}), 400
        if od_duration == 'n_days' and od_days <= 0:
            return jsonify({'success': False, 'message': 'Specify number of days for OD'}), 400
        # from/to are optional for OD in this flow
    elif pass_type == 'other':
        # Hours-based, still expects from/to to bound the time window
        if not other_hours:
            return jsonify({'success': False, 'message': 'Specify hours/time for Other'}), 400
        # from/to optional for Other

    # Fetch student basic info
    conn_local = sqlite3.connect('school.db')
    cur = conn_local.cursor()
    cur.execute("SELECT name, rollno, current_semester FROM students WHERE user_id=?", (session.get('user'),))
    row = cur.fetchone()
    if not row:
        conn_local.close()
        return jsonify({'success': False, 'message': 'Student not found'}), 404
    requester_name, rollno, department = (row[0] or ''), (row[1] or ''), (row[2] or '')
    cur.execute(
        """
        INSERT INTO out_passes (
            user_role, requester_user_id, requester_name, rollno, department,
            pass_type, reason, from_datetime, to_datetime,
            od_duration, od_days, other_hours,
            status, created_at, updated_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'pending', ?, ?)
        """,
        (
            'student', session.get('user'), requester_name, rollno, department,
            pass_type, reason, from_datetime, to_datetime,
            od_duration, od_days, other_hours,
            _now_epoch(), _now_epoch()
        )
    )
    conn_local.commit()
    pass_id = cur.lastrowid
    conn_local.close()
    return jsonify({'success': True, 'id': pass_id})

@app.route('/out_pass/my', methods=['GET'])
@login_required('student')
def list_my_out_passes():
    conn_local = sqlite3.connect('school.db')
    conn_local.row_factory = sqlite3.Row
    cur = conn_local.cursor()
    cur.execute("SELECT * FROM out_passes WHERE requester_user_id=? ORDER BY created_at DESC", (session.get('user'),))
    rows = cur.fetchall()
    conn_local.close()
    return jsonify({'success': True, 'passes': [dict(r) for r in rows]})

def _role_for_approvals() -> str:
    role = (session.get('role') or '').lower()
    return role

@app.route('/out_pass/pending', methods=['GET'])
@login_required_any(('teacher','hod','principal','admin'))
def list_pending_out_passes():
    # Teachers/HOD/Principal can see pending requests. Scope teachers to department.
    approver_role = _role_for_approvals()
    conn_local = sqlite3.connect('school.db')
    conn_local.row_factory = sqlite3.Row
    cur = conn_local.cursor()

    if approver_role == 'teacher':
        # advisor stage: show items awaiting advisor approval
        cur.execute("SELECT department FROM teachers WHERE user_id=?", (session.get('user'),))
        row = cur.fetchone()
        dept = (row[0] or '') if row else ''
        if dept:
            cur.execute("SELECT * FROM out_passes WHERE advisor_status='pending' AND department=? ORDER BY created_at DESC", (dept,))
        else:
            cur.execute("SELECT * FROM out_passes WHERE advisor_status='pending' ORDER BY created_at DESC")
    elif approver_role == 'hod':
        # fetch teacher dept to scope
        cur.execute("SELECT department FROM teachers WHERE user_id=?", (session.get('user'),))
        row = cur.fetchone()
        dept = (row[0] or '') if row else ''
        if dept:
            cur.execute("SELECT * FROM out_passes WHERE advisor_status='approved' AND hod_status='pending' AND department=? ORDER BY created_at DESC", (dept,))
        else:
            cur.execute("SELECT * FROM out_passes WHERE advisor_status='approved' AND hod_status='pending' ORDER BY created_at DESC")
    else:
        # principal or others
        cur.execute("SELECT * FROM out_passes WHERE (advisor_status='pending' OR hod_status='pending') ORDER BY created_at DESC")

    rows = cur.fetchall()
    conn_local.close()
    return jsonify({'success': True, 'passes': [dict(r) for r in rows]})

@app.route('/out_pass/<int:pass_id>/decision', methods=['POST'])
@login_required_any(('teacher','hod','principal','admin'))
def decide_out_pass(pass_id: int):
    data = request.get_json() or {}
    decision = (data.get('decision') or '').strip().lower()  # 'approved' | 'rejected'
    remarks = (data.get('remarks') or '').strip()
    # Optional edits by approver (teacher can tweak time)
    new_from = (data.get('from_datetime') or '').strip()
    new_to = (data.get('to_datetime') or '').strip()
    if decision not in ('approved', 'rejected'):
        return jsonify({'success': False, 'message': 'Invalid decision'}), 400

    conn_local = sqlite3.connect('school.db')
    cur = conn_local.cursor()
    role = _role_for_approvals()
    now = _now_epoch()
    if role == 'teacher':
        # advisor stage
        # Apply optional time edits only if present
        if new_from or new_to:
            # Read existing from/to to keep unchanged when blank
            cur.execute("SELECT from_datetime, to_datetime FROM out_passes WHERE id=?", (pass_id,))
            row = cur.fetchone()
            cur_from = row[0] if row else ''
            cur_to = row[1] if row else ''
            eff_from = new_from or cur_from
            eff_to = new_to or cur_to
            cur.execute("UPDATE out_passes SET from_datetime=?, to_datetime=? WHERE id=?", (eff_from, eff_to, pass_id))
        cur.execute("UPDATE out_passes SET advisor_status=?, advisor_user_id=?, advisor_remarks=?, updated_at=? WHERE id=?",
                   (decision, session.get('user'), remarks, now, pass_id))
        # If rejected, set final status to rejected as well
        if decision == 'rejected':
            cur.execute("UPDATE out_passes SET status='rejected' WHERE id=?", (pass_id,))
        elif decision == 'approved':
            # move to HOD stage
            cur.execute("UPDATE out_passes SET status='pending' WHERE id=?", (pass_id,))
    elif role == 'hod':
        # final stage
        cur.execute("UPDATE out_passes SET hod_status=?, hod_user_id=?, hod_remarks=?, updated_at=? WHERE id=?",
                   (decision, session.get('user'), remarks, now, pass_id))
        cur.execute("UPDATE out_passes SET status=? WHERE id=?", ('approved' if decision=='approved' else 'rejected', pass_id))
    else:
        # principal/admin can override directly final status
        cur.execute("UPDATE out_passes SET status=?, approver_user_id=?, remarks=?, updated_at=? WHERE id=?",
                   (decision, session.get('user'), remarks, now, pass_id))
    conn_local.commit()
    updated = cur.rowcount
    conn_local.close()
    if updated == 0:
        return jsonify({'success': False, 'message': 'Pass not found'}), 404
    return jsonify({'success': True})


# --- Diagnostics ---
@app.route('/debug/students', methods=['GET'])
@login_required('admin')
def debug_students():
    """Debug endpoint to see raw student data"""
    c.execute("SELECT * FROM students LIMIT 3")
    students = c.fetchall()
    c.execute("PRAGMA table_info(students)")
    columns = c.fetchall()
    return jsonify({
        'columns': [col[1] for col in columns],
        'sample_data': students,
        'total_students': len(students)
    })

@app.route('/health', methods=['GET'])
def health():
    creds_exists = os.path.exists(GOOGLE_CREDENTIALS_FILE)
    # Try to infer whether it's a service account without exposing secrets
    sa_detected = False
    service_account_email = None
    try:
        if creds_exists:
            with open(GOOGLE_CREDENTIALS_FILE, 'r') as f:
                data = json.load(f)
            sa_detected = isinstance(data, dict) and data.get('type') == 'service_account'
            if sa_detected:
                service_account_email = data.get('client_email', 'Unknown')
    except Exception:
        sa_detected = False
    
    def mask(value):
        if not value:
            return None
        return value[:6] + '...' + value[-4:]
    
    # Check database status
    try:
        c.execute("SELECT COUNT(*) FROM attendance")
        attendance_count = c.fetchone()[0]
        c.execute("SELECT COUNT(*) FROM students")
        students_count = c.fetchone()[0]
    except Exception as e:
        attendance_count = f"Error: {e}"
        students_count = f"Error: {e}"
    
    return jsonify({
        'success': True,
        'credentials_file': GOOGLE_CREDENTIALS_FILE,
        'credentials_found': creds_exists,
        'service_account_detected': sa_detected,
        'service_account_email': service_account_email,
        'students_sheet_id': mask(STUDENTS_SHEET_ID),
        'attendance_sheet_id': mask(ATTENDANCE_SHEET_ID),
        'students_range': STUDENTS_RANGE,
        'attendance_range': ATTENDANCE_RANGE,
        'excel_mode': USE_EXCEL_ONLY,
        'students_xlsx_found': os.path.exists(STUDENTS_XLSX),
        'attendance_xlsx_found': os.path.exists(ATTENDANCE_XLSX),
        'database_status': {
            'attendance_records': attendance_count,
            'students_records': students_count
        }
    })

# Add this new route to your app.py file

@app.route('/add_teacher', methods=['POST'])
@login_required('admin')
def add_teacher():
    # Get the data sent from the JavaScript form
    data = request.get_json() or {}
    # Frontend sends keys: teacher_name, department, user_id, password, role (optional: teacher|hod|principal)
    teacher_name = (data.get('teacher_name') or data.get('name') or '').strip()
    department = (data.get('department') or '').strip()
    user_id = (data.get('user_id') or '').strip()
    password = (data.get('password') or '').strip()
    role = (data.get('role') or 'teacher').strip().lower()

    # Basic validation to ensure required fields are present
    if not all([teacher_name, department, user_id, password]):
        return jsonify({"success": False, "message": "All fields are required"}), 400

    # Hash the password for security
    hashed_pw = generate_password_hash(password)

    # Detect whether 'role' column exists to avoid SQL errors on older DBs
    def _has_column(table_name: str, column_name: str) -> bool:
        try:
            c.execute(f"PRAGMA table_info({table_name})")
            cols = {row[1] for row in c.fetchall()}
            return column_name in cols
        except Exception:
            return False

    try:
        # ID is AUTOINCREMENT; do not provide it explicitly
        if _has_column('teachers', 'role'):
            c.execute(
                "INSERT INTO teachers (teacher_name, department, user_id, pass_hash, pass_plain, role) VALUES (?, ?, ?, ?, ?, ?)",
                (teacher_name, department, user_id, hashed_pw, password, role)
            )
        else:
            # Fallback for legacy DB schema without 'role'
            c.execute(
                "INSERT INTO teachers (teacher_name, department, user_id, pass_hash, pass_plain) VALUES (?, ?, ?, ?, ?)",
                (teacher_name, department, user_id, hashed_pw, password)
            )
        conn.commit()
        return jsonify({"success": True, "message": "Teacher added successfully"})

    except sqlite3.IntegrityError:
        return jsonify({"success": False, "message": "This Username already exists"}), 409
    except sqlite3.OperationalError as e:
        # Return helpful diagnostics
        return jsonify({"success": False, "message": f"Database error: {str(e)}"}), 500
    except Exception as e:
        return jsonify({"success": False, "message": f"Unexpected error: {str(e)}"}), 500


@app.route("/teachers")
def get_teachers():
    conn = sqlite3.connect("school.db")
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT * FROM teachers")
    rows = cur.fetchall()
    
    teacher_list = []
    for t in rows:
        # Parse extra_json if present
        extra = {}
        if "extra_json" in t.keys() and t["extra_json"]:
            try:
                extra = json.loads(t["extra_json"])
            except Exception:
                extra = {}

        def safe_get(col, fallback=""):
            val = t[col] if col in t.keys() else None
            if val is None or str(val).strip() == "":
                return extra.get(col, fallback)
            return val

        teacher_data = {
            "id": t["id"],
            "teacher_name": safe_get("teacher_name"),
            "department": safe_get("department"),
            "user_id": safe_get("user_id"),
            "password": t["pass_plain"] or "",
            # "email": safe_get("email"),   # <-- REMOVE
            # "phone": safe_get("phone"),   # <-- REMOVE
            "qualification": safe_get("qualification"),
            "experience": safe_get("experience"),
            "subject": safe_get("subject"),
            "address": safe_get("address"),
            "date_of_joining": safe_get("date_of_joining"),
            "salary": safe_get("salary"),
            "extra": extra
        }
        teacher_list.append(teacher_data)
    
    conn.close()
    return jsonify(teacher_list)
    

# === DELETE STUDENT ROUTE ===
@app.route('/delete_student/<int:student_id>', methods=['DELETE'])
@login_required('admin')
def delete_student(student_id):
    c.execute("DELETE FROM students WHERE id=?", (student_id,))
    conn.commit()
    if c.rowcount == 0:
        return jsonify({"success": False, "message": "Student not found"}), 404
    return jsonify({"success": True, "message": "Student deleted"})

@app.route('/delete_teacher/<int:teacher_id>', methods=['DELETE'])
@login_required('admin')
def delete_teacher(teacher_id):
    # Your database logic to delete the teacher by ID
    c.execute("DELETE FROM teachers WHERE id=?", (teacher_id,))
    conn.commit()
    if c.rowcount == 0:
        return jsonify({"success": False, "message": "Teacher not found"}), 404
    return jsonify({"success": True, "message": "Teacher deleted"})
    
# === RESET STUDENT PASSWORD ROUTE ===
@app.route('/reset_student_password/<int:student_id>', methods=['POST'])
@login_required('admin')
def reset_student_password(student_id):
    new_password = generate_password()
    new_password_hash = generate_password_hash(new_password)

    c.execute("UPDATE students SET password_hash=?, password_plain=? WHERE id=?",
              (new_password_hash, new_password, student_id))
    conn.commit()

    if c.rowcount == 0:
        return jsonify({"success": False, "message": "Student not found"}), 404

    return jsonify({"success": True, "new_password": new_password})


@app.route('/reset_teacher_password', methods=['POST'])
@login_required('teacher')
def reset_teacher_password():
    # 1. Ensure teacher is logged in
    username = session.get('user')
    if not username:
        return jsonify({'success': False, 'message': 'User session expired'})

    # 2. Get new password from request
    data = request.get_json()
    new_pass = data.get('password', '').strip()
    if not new_pass:
        return jsonify({'success': False, 'message': 'Password cannot be empty'})

    # 3. Hash the new password
    hashed_password = generate_password_hash(new_pass)

    try:
        # 4. Update only this teacher's password
        c.execute(
            "UPDATE teachers SET pass_hash=?, pass_plain=? WHERE user_id=?",
            (hashed_password, new_pass, username)
        )
        conn.commit()

        if c.rowcount == 0:
            return jsonify({'success': False, 'message': 'Teacher not found'})

        return jsonify({'success': True, 'message': 'Password changed successfully'})
    except Exception as e:
        print("Error updating password:", e)
        return jsonify({'success': False, 'message': 'Error updating password'})



# === STUDENT ATTENDANCE AVERAGE ROUTE ===
@app.route('/student_attendance_average', methods=['GET'])
@login_required('student')
def get_student_attendance_average():
    # Always sync attendance from Google Sheets before calculation
    try:
        load_attendance_from_gsheets()
    except Exception as e:
        print("Attendance sync failed:", e)

    user_id = session.get('user')
    try:
        conn_local = sqlite3.connect('school.db')
        cur = conn_local.cursor()
        cur.execute("SELECT rollno FROM students WHERE user_id=?", (user_id,))
        student = cur.fetchone()
        if not student:
            conn_local.close()
            return jsonify({"success": False, "message": "Student not found"}), 404
        rollno = student[0]
        cur.execute("SELECT status FROM attendance WHERE rollno=?", (rollno,))
        attendance_records = cur.fetchall()
        conn_local.close()
    except Exception as e:
        try:
            conn_local.close()
        except Exception:
            pass
        return jsonify({"success": False, "message": f"Database error: {e}"}), 500
    
    if not attendance_records:
        return jsonify({
            "success": True, 
            "attendance_average": 0, 
            "total_days": 0, 
            "present_days": 0,
            "message": "No attendance records found"
        })
    # Count only non-blank days as working days, and handle various present/absent formats
    def present_status(s):
        if not s or not s.strip():
            return False
        status = s.strip().upper()
        return status in ['P', 'PRESENT', '1', 'YES', 'Y']
    
    def absent_status(s):
        if not s or not s.strip():
            return False
        status = s.strip().upper()
        return status in ['A', 'ABSENT', '0', 'NO', 'N']
    
    def valid_status(s):
        return s and s.strip() != ''
    total_days = sum(1 for record in attendance_records if valid_status(record[0]))
    present_days = sum(1 for record in attendance_records if present_status(record[0]))
    absent_days = sum(1 for record in attendance_records if absent_status(record[0]))
    attendance_average = (present_days / total_days * 100) if total_days > 0 else 0
    return jsonify({
        "success": True,
        "attendance_average": round(attendance_average, 2),
        "total_days": total_days,
        "present_days": present_days,
        "absent_days": absent_days
    })


# === ALL STUDENTS ATTENDANCE AVERAGES ROUTE ===
@app.route('/all_students_attendance_averages', methods=['GET'])
@login_required('admin')
def get_all_students_attendance_averages():
    # Auto-sync attendance before computing
    # Always sync attendance from Google Sheets before calculation
    try:
        load_attendance_from_gsheets()
    except Exception as e:
        print("Attendance sync failed:", e)
    
    # Use a local connection to avoid cursor recursion
    conn_local = sqlite3.connect('school.db')
    cur_local = conn_local.cursor()

    # Get all students (use current_semester instead of student_class)
    cur_local.execute("SELECT id, name, rollno, reg_no, current_semester FROM students")
    students = cur_local.fetchall()
    
    attendance_data = []
    
    for student in students:
        student_id, name, rollno, reg_no, student_class = student
        
        # Get attendance records for this student
        cur_local.execute("SELECT status FROM attendance WHERE rollno=?", (rollno,))
        attendance_records = cur_local.fetchall()
        def present_status(s):
            if not s or not s.strip():
                return False
            status = s.strip().upper()
            return status in ['P', 'PRESENT', '1', 'YES', 'Y']
        
        def absent_status(s):
            if not s or not s.strip():
                return False
            status = s.strip().upper()
            return status in ['A', 'ABSENT', '0', 'NO', 'N']
        
        def valid_status(s):
            return s and s.strip() != ''
        total_days = sum(1 for record in attendance_records if valid_status(record[0]))
        present_days = sum(1 for record in attendance_records if present_status(record[0]))
        absent_days = sum(1 for record in attendance_records if absent_status(record[0]))
        attendance_average = (present_days / total_days * 100) if total_days > 0 else 0
        attendance_data.append({
            "student_id": student_id,
            "name": name,
            "rollno": rollno,
            "reg_no": reg_no,
            "class": student_class,
            "attendance_average": round(attendance_average, 2),
            "total_days": total_days,
            "present_days": present_days,
            "absent_days": absent_days
        })
    
    conn_local.close()
    return jsonify({
        "success": True,
        "students": attendance_data
    })

@app.route('/teacher/all_students_attendance_averages', methods=['GET'])
@login_required('teacher')
def teacher_all_students_attendance_averages():
    # Auto-sync attendance before computing
    # Always sync attendance from Google Sheets before calculation
    try:
        load_attendance_from_gsheets()
    except Exception as e:
        print("Attendance sync failed:", e)
    # Scope to teacher's department
    try:
        conn_local = sqlite3.connect('school.db')
        cur_local = conn_local.cursor()
        cur_local.execute("SELECT department FROM teachers WHERE user_id=?", (session.get('user'),))
        row = cur_local.fetchone()
        conn_local.close()
        dept = (row[0] or '').strip() if row and row[0] else None
    except Exception:
        try:
            conn_local.close()
        except Exception:
            pass
        dept = None
    if dept:
        c.execute("SELECT id, name, rollno, reg_no, current_semester FROM students WHERE current_semester = ?", (dept,))
    else:
        c.execute("SELECT id, name, rollno, reg_no, current_semester FROM students")
    students = c.fetchall()
    
    attendance_data = []
    for student in students:
        student_id, name, rollno, reg_no, student_class = student
        c.execute("SELECT status FROM attendance WHERE rollno=?", (rollno,))
        attendance_records = c.fetchall()
        def present_status(s):
            if not s or not s.strip():
                return False
            status = s.strip().upper()
            return status in ['P', 'PRESENT', '1', 'YES', 'Y']
        
        def absent_status(s):
            if not s or not s.strip():
                return False
            status = s.strip().upper()
            return status in ['A', 'ABSENT', '0', 'NO', 'N']
        
        def valid_status(s):
            return s and s.strip() != ''
        total_days = sum(1 for record in attendance_records if valid_status(record[0]))
        present_days = sum(1 for record in attendance_records if present_status(record[0]))
        absent_days = sum(1 for record in attendance_records if absent_status(record[0]))
        attendance_average = (present_days / total_days * 100) if total_days > 0 else 0
        attendance_data.append({
            "student_id": student_id,
            "name": name,
            "rollno": rollno,
            "reg_no": reg_no,
            "class": student_class,
            "attendance_average": round(attendance_average, 2),
            "total_days": total_days,
            "present_days": present_days,
            "absent_days": absent_days
        })
    
    return jsonify({
        "success": True,
        "students": attendance_data
    })

@app.route('/teacher/daily_absent_students', methods=['GET'])
@login_required('teacher')
def teacher_daily_absent_students():
    # Auto-sync attendance before checking
    try:
        if int(time.time()) - _last_attendance_sync_ts > SYNC_TTL_SECONDS:
            if USE_EXCEL_ONLY or not ATTENDANCE_SHEET_ID:
                load_attendance_from_excel()
            else:
                load_attendance_from_gsheets()
    except Exception as e:
        print("Attendance auto-sync failed:", e)

    # Scope to teacher's department
    try:
        conn_local = sqlite3.connect('school.db')
        cur_local = conn_local.cursor()
        cur_local.execute("SELECT department FROM teachers WHERE user_id=?", (session.get('user'),))
        row = cur_local.fetchone()
        dept = (row[0] or '').strip() if row and row[0] else None
    except Exception:
        try:
            conn_local.close()
        except Exception:
            pass
        dept = None

    absent_students = []
    try:
        if 'conn_local' not in locals():
            conn_local = sqlite3.connect('school.db')
            cur_local = conn_local.cursor()
        if dept:
            cur_local.execute("SELECT rollno, name, current_semester FROM students WHERE current_semester=?", (dept,))
        else:
            cur_local.execute("SELECT rollno, name, current_semester FROM students")
        students = cur_local.fetchall()

        today_iso = time.strftime('%Y-%m-%d').lower()
        today_dmy = time.strftime('%d-%m-%Y').lower()
        today_dmy_mon = time.strftime('%d-%b-%Y').lower()
        today_dmy_mon2 = time.strftime('%d-%b-%y').lower()
        for rollno, name, current_semester in students:
            c.execute("SELECT status FROM attendance WHERE rollno=? AND LOWER(date) IN (?, ?, ?, ?)", (rollno, today_iso, today_dmy, today_dmy_mon, today_dmy_mon2))
            attendance_record = c.fetchone()
            if not attendance_record or (attendance_record[0] and attendance_record[0].lower() in ['absent', 'a', '0', 'no']):
                absent_students.append({
                    "rollno": rollno,
                    "name": name,
                    "class": current_semester,
                    "status": "Absent" if attendance_record else "No record"
                })
        conn_local.close()
    except Exception as e:
        print("Error fetching teacher daily absent:", e)
        try:
            conn_local.close()
        except Exception:
            pass
        return jsonify({"success": False, "message": "Error fetching daily absent students"}), 500

    return jsonify({"success": True, "absent_students": absent_students})

@app.route('/hod/all_students_attendance_averages', methods=['GET'])
@login_required('hod')
def hod_all_students_attendance_averages():
    # Auto-sync attendance before computing
    # Always sync attendance from Google Sheets before calculation
    try:
        load_attendance_from_gsheets()
    except Exception as e:
        print("Attendance sync failed:", e)
    
    # Use a local connection to avoid cursor recursion
    conn_local = sqlite3.connect('school.db')
    cur_local = conn_local.cursor()

    # Return all students with all fields
    cur_local.execute("SELECT * FROM students")
    students = cur_local.fetchall()
    columns = [desc[0] for desc in cur_local.description]
    attendance_data = []
    for student in students:
        student_dict = dict(zip(columns, student))
        # Parse extra_json if present
        extra = {}
        if student_dict.get("extra_json"):
            try:
                extra = json.loads(student_dict["extra_json"])
            except Exception:
                extra = {}
        # Merge extra fields, but don't overwrite main columns
        for k, v in extra.items():
            if k not in student_dict or not student_dict[k]:
                student_dict[k] = v
        # Attendance calculation
        rollno = student_dict.get("rollno", "")
        cur_local.execute("SELECT status FROM attendance WHERE rollno= ?", (rollno,))
        attendance_records = cur_local.fetchall()
        def present_status(s):
            if not s or not s.strip():
                return False
            status = s.strip().upper()
            return status in ['P', 'PRESENT', '1', 'YES', 'Y']
        
        def absent_status(s):
            if not s or not s.strip():
                return False
            status = s.strip().upper()
            return status in ['A', 'ABSENT', '0', 'NO', 'N']
        
        def valid_status(s):
            return s and s.strip() != ''
        total_days = sum(1 for record in attendance_records if valid_status(record[0]))
        present_days = sum(1 for record in attendance_records if present_status(record[0]))
        absent_days = sum(1 for record in attendance_records if absent_status(record[0]))
        attendance_average = (present_days / total_days * 100) if total_days > 0 else 0
        # Add attendance stats to student_dict
        student_dict["attendance_average"] = round(attendance_average, 2)
        student_dict["total_days"] = total_days
        student_dict["present_days"] = present_days
        student_dict["absent_days"] = absent_days
        # Remove sensitive fields
        student_dict.pop("password_hash", None)
        student_dict.pop("password_plain", None)
        student_dict.pop("extra_json", None)
        attendance_data.append(student_dict)
    
    conn_local.close()
    return jsonify({"success": True, "students": attendance_data})

@app.route('/hod/daily_absent_students', methods=['GET'])
@login_required('hod')
def hod_daily_absent_students():
    # Auto-sync attendance before checking
    try:
        if int(time.time()) - _last_attendance_sync_ts > SYNC_TTL_SECONDS:
            if USE_EXCEL_ONLY or not ATTENDANCE_SHEET_ID:
                load_attendance_from_excel()
            else:
                load_attendance_from_gsheets()
    except Exception as e:
        print("Attendance auto-sync failed:", e)

    # Get HOD's department using a local connection
    try:
        conn_local = sqlite3.connect('school.db')
        cur_local = conn_local.cursor()
        cur_local.execute("SELECT department FROM teachers WHERE user_id=?", (session.get('user'),))
        row = cur_local.fetchone()
        department = (row[0] or '').strip() if row and row[0] else None
    except Exception:
        department = None

    # Collect absent students for the department
    absent_students = []
    try:
        if 'conn_local' not in locals():
            conn_local = sqlite3.connect('school.db')
            cur_local = conn_local.cursor()
        if department:
            cur_local.execute("SELECT rollno, name, current_semester FROM students WHERE current_semester=?", (department,))
        else:
            cur_local.execute("SELECT rollno, name, current_semester FROM students")
        students = cur_local.fetchall()

        date_variants = _get_target_date_variants_for_attendance()
        for rollno, name, current_semester in students:
            cur_local.execute(
                "SELECT status FROM attendance WHERE rollno=? AND LOWER(date) IN (?, ?, ?, ?)",
                (rollno, *(date_variants + ['',''])[:4])
            )
            attendance_record = cur_local.fetchone()
            if not attendance_record or (attendance_record[0] and attendance_record[0].lower() in ['absent', 'a', '0', 'no']):
                absent_students.append({
                    "rollno": rollno,
                    "name": name,
                    "class": current_semester,
                    "status": "Absent" if attendance_record else "No record"
                })
        conn_local.close()
    except Exception as e:
        print("Error fetching daily absent students:", e)
        try:
            conn_local.close()
        except Exception:
            pass
        return jsonify({"success": False, "message": "Error fetching daily absent students"}), 500

    return jsonify({"success": True, "absent_students": absent_students})

# --- Principal: daily absent (college-wide) ---
@app.route('/principal/daily_absent_students', methods=['GET'])
@login_required('principal')
def principal_daily_absent_students():
    # Auto-sync attendance before checking
    try:
        if int(time.time()) - _last_attendance_sync_ts > SYNC_TTL_SECONDS:
            if USE_EXCEL_ONLY or not ATTENDANCE_SHEET_ID:
                load_attendance_from_excel()
            else:
                load_attendance_from_gsheets()
    except Exception as e:
        print("Attendance auto-sync failed:", e)

    absent_students = []
    try:
        conn_local = sqlite3.connect('school.db')
        cur_local = conn_local.cursor()
        cur_local.execute("SELECT rollno, name, current_semester FROM students")
        students = cur_local.fetchall()
        date_variants = _get_target_date_variants_for_attendance()
        for rollno, name, current_semester in students:
            cur_local.execute(
                "SELECT status FROM attendance WHERE rollno=? AND LOWER(date) IN (?, ?, ?, ?)",
                (rollno, *(date_variants + ['',''])[:4])
            )
            attendance_record = cur_local.fetchone()
            if not attendance_record or (attendance_record[0] and attendance_record[0].lower() in ['absent', 'a', '0', 'no']):
                absent_students.append({
                    "rollno": rollno,
                    "name": name,
                    "class": current_semester,
                    "status": "Absent" if attendance_record else "No record"
                })
        conn_local.close()
    except Exception as e:
        print("Error fetching principal daily absent:", e)
        try:
            conn_local.close()
        except Exception:
            pass
        return jsonify({"success": False, "message": "Error fetching daily absent students"}), 500

    return jsonify({"success": True, "absent_students": absent_students})

# --- Admin: daily absent (college-wide) ---
@app.route('/daily_absent_students', methods=['GET'])
@login_required('admin')
def admin_daily_absent_students():
    try:
        if int(time.time()) - _last_attendance_sync_ts > SYNC_TTL_SECONDS:
            if USE_EXCEL_ONLY or not ATTENDANCE_SHEET_ID:
                load_attendance_from_excel()
            else:
                load_attendance_from_gsheets()
    except Exception as e:
        print("Attendance auto-sync failed:", e)

    absent_students = []
    try:
        c.execute("SELECT rollno, name, current_semester FROM students")
        students = c.fetchall()
        date_variants = _get_target_date_variants_for_attendance()
        for rollno, name, current_semester in students:
            c.execute(
                "SELECT status FROM attendance WHERE rollno=? AND LOWER(date) IN (?, ?, ?, ?)",
                (rollno, *(date_variants + ['',''])[:4])
            )
            attendance_record = c.fetchone()
            if not attendance_record or (attendance_record[0] and attendance_record[0].lower() in ['absent', 'a', '0', 'no']):
                absent_students.append({
                    "rollno": rollno,
                    "name": name,
                    "class": current_semester,
                    "status": "Absent" if attendance_record else "No record"
                })
    except Exception as e:
        print("Error fetching admin daily absent:", e)
        return jsonify({"success": False, "message": "Error fetching daily absent students"}), 500

    return jsonify({"success": True, "absent_students": absent_students, "total_absent": len(absent_students)})

# --- Principal: attendance averages (college-wide) ---
@app.route('/principal/all_students_attendance_averages', methods=['GET'])
@login_required('principal')
def principal_all_students_attendance_averages():
       # Always sync attendance from Google Sheets before calculation
    try:
        load_attendance_from_gsheets()
    except Exception as e:
        pass

    conn_local = sqlite3.connect('school.db')
    cur_local = conn_local.cursor()

    cur_local.execute("SELECT id, name, rollno, reg_no, current_semester FROM students")
    students = cur_local.fetchall()
    attendance_data = []
    for student in students:
        student_id, name, rollno, reg_no, student_class = student
        cur_local.execute("SELECT status FROM attendance WHERE rollno=?", (rollno,))
        attendance_records = cur_local.fetchall()
        def present_status(s):
            if not s or not s.strip():
                return False
            status = s.strip().upper()
            return status in ['P', 'PRESENT', '1', 'YES', 'Y']
        
        def absent_status(s):
            if not s or not s.strip():
                return False
            status = s.strip().upper()
            return status in ['A', 'ABSENT', '0', 'NO', 'N']
        
        def valid_status(s):
            return s and s.strip() != ''
        total_days = sum(1 for record in attendance_records if valid_status(record[0]))
        present_days = sum(1 for record in attendance_records if present_status(record[0]))
        absent_days = sum(1 for record in attendance_records if absent_status(record[0]))
        attendance_average = (present_days / total_days * 100) if total_days > 0 else 0
        attendance_data.append({
            "student_id": student_id,
            "name": name,
            "rollno": rollno,
            "reg_no": reg_no,
            "class": student_class,
            "attendance_average": round(attendance_average, 2),
            "total_days": total_days,
            "present_days": present_days,
            "absent_days": absent_days
        })
    
    conn_local.close()
    return jsonify({"success": True, "students": attendance_data})
 
# === DEBUG: Show all attendance status for a given roll number ===

# --- RUN SERVER ---
if __name__ == '__main__':
    # Startup summary
    try:
        c.execute("SELECT COUNT(*) FROM students")
        stu_count = c.fetchone()[0]
    except Exception:
        stu_count = 0
    try:
        c.execute("SELECT COUNT(*) FROM attendance")
        att_count = c.fetchone()[0]
    except Exception:
        att_count = 0
    print(f"Startup summary → Students: {stu_count}, Attendance rows: {att_count}, Excel mode: {USE_EXCEL_ONLY}")
    app.run(debug=True)
